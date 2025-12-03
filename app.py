from flask import Flask, render_template, request, redirect, url_for, flash, send_file,jsonify
from flask_sqlalchemy import SQLAlchemy
import random
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
import os
import csv
import time
from io import TextIOWrapper, BytesIO
from PyPDF2 import PdfMerger
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from pathlib import Path
from zipfile import ZipFile
from flask import Response
from io import StringIO
from sqlalchemy.dialects.sqlite import JSON
from sqlalchemy import or_
from sqlalchemy import asc


app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///timetable.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'your_secret_key_here'
db = SQLAlchemy(app)

# Database Models
class Department(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)

class Classroom(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    building_id = db.Column(db.Integer, db.ForeignKey('building.id'), nullable=False)

class Course(db.Model):
    __table_args__ = (
        db.UniqueConstraint('name', 'batch_id', name='unique_course_per_batch'),
    )
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), nullable=False)
    lecture_hour = db.Column(db.Integer, default=0, nullable=False)   # default added
    lab_hour = db.Column(db.Integer, default=2, nullable=False)       # default added
    tutorial_hour = db.Column(db.Integer, default=0, nullable=False)  # default added
    course_code = db.Column(db.String(20), default="CSE", nullable=False)
    is_lab = db.Column(db.Boolean, default=False)
    priority = db.Column(db.Boolean, default=False)# for 2 hour class
    priority_morning = db.Column(db.Boolean, default=False)# for morning shift
    priority_evening = db.Column(db.Boolean, default=False)# for evening shif
    avoid_day = db.Column(db.Integer, default=-1)# to avoid any day
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=True)
    lab_professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=True)
    lab_id1 = db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)
    lab_id2= db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)
    lab_id3 = db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)
    batch_id= db.Column(db.Integer, db.ForeignKey('batch.id'),nullable=True)
    lab_priority_morning = db.Column(db.Boolean, default=False)# for morning shift
    lab_priority_evening = db.Column(db.Boolean, default=False)# for evening shif
    divide = db.Column(db.Boolean, default=False)
    tutorial = db.Column(db.Boolean, default=False)

    priority_classroom_1 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)
    priority_classroom_2 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)
    priority_classroom_3 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)

class Building(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)
    # Relationship: One building → many classrooms
    classrooms = db.relationship('Classroom', backref='building', lazy=True)

class Professor(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)

    # Optional: relationship for easy access
    department = db.relationship('Department', backref=db.backref('professors', lazy=True))
    #email = db.Column(db.String(100), nullable=False, unique=True)

    priority_classroom_1 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)
    priority_classroom_2 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)
    priority_classroom_3 = db.Column(db.Integer, db.ForeignKey('building.id'),default=-1, nullable=True)
    #department = db.Column(db.String(50), nullable=False)
    #designation = db.Column(db.String(50))  # e.g., Lecturer, Assistant Prof, etc.
    #max_hours_per_day = db.Column(db.Integer, default=6)  # for timetable allocation

    # Optional: courses assigned to this professor (many-to-many)

class Lab(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)

class Batch(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    courses = db.relationship('Course', backref='batch', lazy=True)
    odd_sem = db.Column(db.Boolean, default=False,nullable=False)
    department_id = db.Column(db.Integer, db.ForeignKey('department.id'), nullable=False)
    semester = db.Column(db.Integer, nullable=False)

    # Optional: relationship for easy access
    department = db.relationship('Department', backref=db.backref('batches', lazy=True))

class CombinedCourse(db.Model):
    id = db.Column(db.Integer, primary_key=True)

    # Basic fields
    name = db.Column(db.String(150), nullable=False)
    lecture = db.Column(db.Integer, nullable=False)
    capacity = db.Column(db.Integer, nullable=False)
    is_odd = db.Column(db.Boolean, default=False)
    professor_id = db.Column(db.Integer, db.ForeignKey("professor.id"), nullable=False)

    # Stores list of batches + their course codes
    # Example:
    # [
    #   {"batch_id": 1, "course_code": "CS101"},
    #   {"batch_id": 3, "course_code": "IT107"}
    # ]
    batch_course_list = db.Column(JSON, nullable=False)

class Schedule(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    batch_id = db.Column(db.Integer, db.ForeignKey('batch.id'), nullable=False)
    classroom_id = db.Column(db.Integer, db.ForeignKey('classroom.id'), nullable=True)
    lab_id = db.Column(db.Integer, db.ForeignKey('lab.id'), nullable=True)
    course_id = db.Column(db.Integer, db.ForeignKey('course.id'), default=None, nullable=True)
    combined_course_id = db.Column(db.Integer, db.ForeignKey('combined_course.id'), default=None, nullable=True)
    elective_course_id = db.Column(db.Integer, default=None, nullable=True)
    professor_id = db.Column(db.Integer, db.ForeignKey('professor.id'), nullable=True)
    day = db.Column(db.Integer, nullable=False)  # 0-4 for Monday-Friday
    slot = db.Column(db.Integer, nullable=False)  # 0-8 for time slots
    semester = db.Column(db.Boolean, nullable=False)  # 0-8 for time slots
    tutorial = db.Column(db.Boolean,default=False, nullable=True)
    divide_id=db.Column(db.Integer,default=0,nullable=True)
    elective_id=db.Column(db.Integer,default=0,nullable=True)
    name = db.Column(db.String(150),default="", nullable=False)

class Elective(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    department_id = db.Column(db.Integer, nullable=False)
    semester = db.Column(db.Integer, nullable=False)
    names = db.Column(db.Text, nullable=False)  # stores "A,B,C"
    professor_id = db.Column(db.Text, nullable=False)
    lab_id = db.Column(db.Text, nullable=False)
    lab_professor_id = db.Column(db.Text, nullable=False)
    lecture_hour = db.Column(db.Integer, nullable=False)
    lab_hour = db.Column(db.Integer, nullable=False)
    tutorial_hour = db.Column(db.Integer, nullable=False)


def find_available_classroom_with_priorityroom(day, slot,classroom_id,batch,sem=False):
    classrooms = Classroom.query.filter_by(building_id=classroom_id)
    for classroom in classrooms:
        if batch.capacity > classroom.capacity:
            continue
        slot_occupied = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot ,semester=sem
        ).first()
        slot_occupied2 = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot+1,semester=sem
        ).first()
        if not (slot_occupied or slot_occupied2):
            return classroom  # classroom is free for both slots
    return None

def find_available_classroom(day, slot , batch,sem=False):
    candidates = (
        Classroom.query
        .filter(Classroom.capacity >= batch.capacity)
        .order_by(asc(Classroom.capacity))
        .all()
    )

    for classroom in candidates:
        occupied = Schedule.query.filter_by(
            classroom_id=classroom.id,
            day=day,
            slot=slot,
            semester=sem
        ).first()
        occupied2 = Schedule.query.filter_by(
            classroom_id=classroom.id,
            day=day,
            slot=slot+1,
            semester=sem
        ).first()
        if occupied or occupied2:
            continue
        return classroom

    return None

def find_available_classroom_with_priorityroom_onehour(day,slot,classroom_id,batch,sem=False):
    classrooms = Classroom.query.filter_by(building_id=classroom_id)
    for classroom in classrooms:
        if batch.capacity > classroom.capacity:
            continue
        slot_occupied = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot,semester=sem
        ).first()
        if not slot_occupied:
            return classroom  # classroom is free for both slots
    return None

def find_combined_classroom(day,slot,sem,capacity):
    all_classrooms = Classroom.query.all()
    for classroom in all_classrooms:
        if classroom.capacity<capacity:
            continue
        slot_occupied = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot,semester=sem
        ).first()

        if not slot_occupied:
            return classroom  # classroom is free for both slots
    return None

def find_available_classroom_onehour(day,slot,batch,sem):
    candidates = (
        Classroom.query
        .filter(Classroom.capacity >= batch.capacity)
        .order_by(asc(Classroom.capacity))
        .all()
    )
    for classroom in candidates:
        occupied = Schedule.query.filter_by(
            classroom_id=classroom.id,
            day=day,
            slot=slot,
            semester=sem
        ).first()
        if occupied:
            continue
        return classroom

    return None

def find_available_lab(day, slot,id,batch,sem,hour):
    all_labs=[]
    if not batch:
        return False
    if id==-1:
        all_labs = Lab.query.all()
    else:
        all_labs = Lab.query.filter_by(id=id).all()
    for lab in all_labs:
        ans=True
        for i in range(hour):
            slot_occupied = Schedule.query.filter_by(
                lab_id=lab.id, day=day, slot=slot+i,semester=sem
            ).first()
            if slot_occupied:
                ans=False
        if ans:
            return lab  # classroom is free for both slots
    return None

def is_slot_available(course, day, slot, building_id,batch,sem=False):
    if not batch:
        return False
    existing_schedule = Schedule.query.filter_by(
        batch_id=course["batch_id"], day=day, slot=slot,semester=sem
    ).first()
    professor_schedule = Schedule.query.filter_by(
        professor_id=course["professor_id"], day=day, slot=slot,semester=sem
    ).first()
    if existing_schedule or professor_schedule:
        return False
    classrooms=None
    if(building_id==-1):
        classrooms=Classroom.query.all()
    else :
        classrooms=Classroom.query.filter_by(building_id=building_id).all()

    for classroom in classrooms:
        if batch.capacity > classroom.capacity:
            continue
        classroom_schedule = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot,semester=sem
        ).first()

        if not classroom_schedule:
            return True
    return False

def is_slot_available_lab_priority1(course, day, slot,batch,sem,idd):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    lab=Lab.query.filter_by(id=course.lab_id1).first()
    if not lab:
        return False
    if not batch:
        return False
    
    for i in range(course.lab_hour):
        existing_schedule = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=idd
        ).first()
        second = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=0
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=course.lab_professor_id, day=day, slot=slot,semester=sem
        ).first()
        lab_schedule = Schedule.query.filter_by(
            lab_id=course.lab_id1, day=day, slot=slot,semester=sem
        ).first()
        if existing_schedule or professor_schedule or lab_schedule or second:
            return False

    return True

def is_slot_available_lab_priority2(course, day, slot,batch,sem,idd):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    lab=Lab.query.filter_by(id=course.lab_id2).first()
    if not lab:
        return False
    if not batch:
        return False
    
    for i in range(course.lab_hour):
        existing_schedule = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=idd
        ).first()
        second = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=0
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=course.lab_professor_id, day=day, slot=slot,semester=sem
        ).first()
        lab_schedule = Schedule.query.filter_by(
            lab_id=course.lab_id2, day=day, slot=slot,semester=sem
        ).first()
        if existing_schedule or professor_schedule or lab_schedule or second:
            return False

    return True

def is_slot_available_lab_priority3(course, day, slot,batch,sem,idd):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    lab=Lab.query.filter_by(id=course.lab_id3).first()
    if not lab:
        return False
    if not batch:
        return False
    
    for i in range(course.lab_hour):
        existing_schedule = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=idd
        ).first()
        second = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=0
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=course.lab_professor_id, day=day, slot=slot,semester=sem
        ).first()
        lab_schedule = Schedule.query.filter_by(
            lab_id=course.lab_id1, day=day, slot=slot,semester=sem
        ).first()
        if existing_schedule or professor_schedule or lab_schedule or second:
            return False

    return True

def is_slot_available_lab(course, day, slot,batch,sem,idd):
    for i in range(course.lab_hour):
        existing_schedule = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=idd
        ).first()
        second = Schedule.query.filter_by(
            batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=0
        ).first()
        professor_schedule = Schedule.query.filter_by(
            professor_id=course.lab_professor_id, day=day, slot=slot,semester=sem
        ).first()
        if existing_schedule or professor_schedule or second:
            return False

    return True

def is_tutorial_slot_available(course, day, slot,batch,divide_id,sem=False):
    if not batch:
        return False
    existing_schedule = Schedule.query.filter_by(
        batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=divide_id
    ).first()
    if existing_schedule:
        return False
    existing_schedules = Schedule.query.filter_by(
        batch_id=course.batch_id, day=day, slot=slot,semester=sem,divide_id=0
    ).first()
    if existing_schedules:
        return False
    classrooms = Classroom.query.order_by(Classroom.capacity.asc()).all()

    for classroom in classrooms:
        if course.divide:
            if int(batch.capacity/2) > classroom.capacity:
                continue
        else:
            if batch.capacity > classroom.capacity:
                continue
        classroom_schedule = Schedule.query.filter_by(
            classroom_id=classroom.id, day=day, slot=slot,semester=sem
        ).first()

        if not classroom_schedule:
            return True
    return False

def is_combined_available(professor,batch_ids,day,slot,sem):
    for id in batch_ids:
        sch=Schedule.query.filter_by(batch_id=id,day=day,slot=slot,semester=sem).first()
        if sch:
            return False
    sch=Schedule.query.filter_by(professor_id=professor.id,day=day,slot=slot,semester=sem).first()
    if sch:
        return False
    return True

def generate_excel_professor(prof_id):
    if not prof_id:
        return None

    prof = Professor.query.filter_by(id=prof_id).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(professor_id=prof.id).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.filter_by(id=schedule.course_id).first()
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.filter_by(id=schedule.professor_id).first()
        classroom = Classroom.query.filter_by(id=schedule.classroom_id).first()
        lab = Lab.query.filter_by(id=schedule.lab_id).first()
        batch=Batch.query.filter_by(id=schedule.batch_id).first()
        
        entry = f"{course.name} ({professor.name})"
        if lab:
            entry += f"[{lab.name}] (P)"
        if classroom:
            entry += f" {{{classroom.name}}} {batch.name} (L)"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Professor: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"professor_{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

def generate_excel(batch_ids):
    if not batch_ids:
        return None
    
    print("WRONG")
    batch = Batch.query.filter_by(id=batch_ids[0]).first()
    if not batch:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(batch_id=batch.id).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.filter_by(id=schedule.course_id).first()
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.filter_by(id=schedule.professor_id).first()
        classroom = Classroom.query.filter_by(id=schedule.classroom_id).first()
        lab = Lab.query.filter_by(id=schedule.lab_id).first()
        
        entry = f"{course.name} ({professor.name})"
        if lab:
            entry += f" [{lab.name}] (P)"
        if classroom:
            entry += f" {{{classroom.name}}}"
            if schedule.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Batch: {batch.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"batch_{batch.name}.xlsx"
    wb.save(excel_path)
    
    return excel_path

def generate_excel_lab(prof_id):
    if not prof_id:
        return None

    prof = Lab.query.filter_by(id=prof_id).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(lab_id=prof.id).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.filter_by(id=schedule.course_id).first()
        professor = Professor.query.filter_by(id=schedule.professor_id).first()
        classroom = Classroom.query.filter_by(id=schedule.classroom_id).first()
        lab = Lab.query.filter_by(id=schedule.lab_id).first()
        
        entry = f"{course.name} ({professor.name})"
        if lab:
            entry += f" [{lab.name}]"
        if classroom:
            entry += f" {{{classroom.name}}}"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Lab: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"lab_{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

def generate_excel_classroom(prof_id):
    if not prof_id:
        return None

    prof = Classroom.query.filter_by(id=prof_id).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(classroom_id=prof.id).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.filter_by(id=schedule.course_id).first()
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.filter_by(id=schedule.professor_id).first()
        classroom = Classroom.query.filter_by(id=schedule.classroom_id).first()
        lab = Lab.query.filter_by(id=schedule.lab_id).first()
        
        entry = f"{course.name} ({professor.name})"
        if lab:
            entry += f" [{lab.name}]"
        if classroom:
            entry += f" {{{classroom.name}}}"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Classroom: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"classroom_{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

def generate_excel_all_batches(batch_ids):
    if not batch_ids:
        return None

    batch = Batch.query.filter_by(id=batch_ids[0]).first()

    if not batch:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(batch_id=batch.id).all()

    for schedule in schedules:
        if schedule.slot == 5:
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        
        entry = ""
        if course:
            entry+=f"{course.name} "
        if professor:
            entry+=f"({professor.name}) "
        if lab:
            entry += f" [{lab.name}] (P )"
        if classroom:
            entry += f" {{{classroom.name}}}"
            if schedule.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Batch: {batch.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"{batch.name}.xlsx"
    wb.save(excel_path)
    
    return excel_path

def generate_excel_all_professors(prof_id,sem):
    if not prof_id:
        return None

    prof = Professor.query.filter_by(id=prof_id[0]).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(professor_id=prof.id,semester=sem).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        batch=Batch.query.filter_by(id=schedule.batch_id).first()
        
        entry = ""
        if course:
            entry+=f"{course.name} "
        if professor:
            entry+=f"({professor.name}) "
        if lab:
            entry += f" [{lab.name}] (P)"
        if classroom:
            entry += f" {{{classroom.name}}} {batch.name} (L)"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Professor: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

def generate_excel_all_classrooms(prof_id,sem):
    if not prof_id:
        return None

    prof = Classroom.query.filter_by(id=prof_id[0]).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(classroom_id=prof.id,semester=sem).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.get(schedule.combined_course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        
        entry = ""
        if course:
            entry+=f"{course.name} "
        if professor:
            entry+=f"({professor.name}) "
        if lab:
            entry += f" [{lab.name}]"
        if classroom:
            entry += f" {{{classroom.name}}}"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Classroom: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

def generate_excel_all_labs(prof_id,sem):
    if not prof_id:
        return None

    prof = Lab.query.filter_by(id=prof_id[0]).first()
    if not prof:
        return None

    time_slots = [
        "08:00 AM - 09:00 AM", "09:00 AM - 10:00 AM", "10:00 AM - 11:00 AM",
        "11:00 AM - 12:00 PM", "12:00 PM - 01:00 PM", "01:00 PM - 02:00 PM",
        "02:00 PM - 03:00 PM", "03:00 PM - 04:00 PM", "04:00 PM - 05:00 PM",
        "05:00 PM - 06:00 PM"
    ]
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

    timetable = [["" for _ in range(10)] for _ in range(5)]
    
    # Set lunch break for all days at slot 4 (12:00 PM - 1:00 PM)
    # for day in range(5):
    timetable[0][5] = "Lunch"
    timetable[1][5] = "Lunch"
    timetable[2][5] = "Lunch"
    timetable[3][5] = "Lunch"
    timetable[4][5] = "Lunch"
    schedules = Schedule.query.filter_by(lab_id=prof.id,semester=sem).all()

    for schedule in schedules:
        if schedule.slot == 5:  # Skip lunch slot
            continue
        course = Course.query.get(schedule.course_id)
        professor = Professor.query.get(schedule.professor_id)
        classroom = Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        
        entry = ""
        if course:
            entry+=f"{course.name} "
        if professor:
            entry+=f"({professor.name}) "
        if lab:
            entry += f" [{lab.name}]"
        if classroom:
            entry += f" {{{classroom.name}}}"
            
        timetable[schedule.day][schedule.slot] = entry

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"TimeTable"

    ws['A1'] = f"Timetable for Lab: {prof.name}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:J1')

    ws['A2'] = "Day"
    for col_index, time_slot in enumerate(time_slots, start=2):
        col_letter = get_column_letter(col_index)
        ws[f'{col_letter}2'] = time_slot

    for day_index, day_name in enumerate(days):
        ws[f"A{day_index + 3}"] = day_name
        for slot_index, slot_content in enumerate(timetable[day_index]):
            col_letter = get_column_letter(slot_index + 2)
            ws[f"{col_letter}{day_index + 3}"] = slot_content

    current_dir = Path(__file__).parent
    timetables_dir = current_dir / "timetables"
    timetables_dir.mkdir(exist_ok=True)
    excel_path = timetables_dir / f"{prof.name}.xlsx"
    wb.save(excel_path)
    print("CORRECT")
    
    return excel_path

#################      Major Functions     ##########

# access as course.lab
def assign_Lab(course,idd,sem=False):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    morning_labp1 = []
    morning_labp2 = []
    morning_labp3 = []
    other=[]

    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(10):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority1(course, day, slot+i,batch,sem,idd)
            if done:
                morning_labp1.append((day, slot))
                
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(10):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority2(course, day, slot+i,batch,sem,idd)
            if done:
                morning_labp2.append((day, slot))
                
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(10):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority3(course, day, slot+i,batch,sem,idd)
            if done:
                morning_labp3.append((day, slot))
                

    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(10):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab(course, day, slot+i,batch,sem,idd)
            if done:
                other.append((day, slot))
                


    day = None
    start_slot = None
    lab_assigned=None
    if morning_labp1:
        day,start_slot = random.choice(morning_labp1)
        lab_assigned=course.lab_id1
    elif morning_labp2:
        day,start_slot = random.choice(morning_labp2)
        lab_assigned=course.lab_id2
    elif morning_labp3:
        day,start_slot = random.choice(morning_labp3)
        lab_assigned=course.lab_id3
    elif other:
        day,start_slot=random.choice(other)
        labsss=find_available_lab(day,start_slot,-1, batch,sem,course.lab_hour)
        lab_assigned=labsss.id
    else:
        print(f"No available slots for course {course.name}")
        return
    try:
        for offset in range(course.lab_hour):
            new_schedule = Schedule(
                batch_id=course.batch_id,
                course_id=course.id,
                professor_id = course.lab_professor_id if course.lab_professor_id else -1,
                lab_id=lab_assigned,
                day=day,
                slot=start_slot + offset,
                classroom_id= None,
                semester=sem,
                divide_id=idd
            )
            db.session.add(new_schedule)
        print("DONNNN")
        db.session.commit()
        print("SUCCESS")
        flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()  # full error trace
        print("Error:", e)

def assign_Evening_lab(course,idd,sem=False):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    morning_labp1 = []
    evening_labp1 = []
    morning_labp2 = []
    evening_labp2 = []
    morning_labp3 = []
    evening_labp3 = []
    other=[]

    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority1(course, day, slot+i,batch,sem,idd)
            if done:
                morning_labp1.append((day, slot))
                
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(6,9):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority1(course, day, slot+i,batch,sem,idd)
            if done:
                evening_labp1.append((day, slot))
                
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=0:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority2(course, day, slot+i,batch,sem,idd)
                if done:
                    morning_labp2.append((day, slot))
                    
    
    for day in range(5):
            if day == course.avoid_day:
                continue 
            for slot in range(6,9):
                if slot==0 or slot==2 or slot==5 or slot==7:
                    continue
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority2(course, day, slot+i,batch,sem,idd)
                if done:
                    evening_labp2.append((day, slot))
                    
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=0:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority3(course, day, slot+i,batch,sem,idd)
                if done:
                    morning_labp3.append((day, slot))
                    
    
    for day in range(5):
            if day == course.avoid_day:
                continue 
            for slot in range(6,9):
                if slot==0 or slot==2 or slot==5 or slot==7:
                    continue
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority3(course, day, slot+i,batch,sem,idd)
                if done:
                    evening_labp3.append((day, slot))
                    

    for day in range(5): 
        for slot in range(9): 
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=5:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab(course, day, slot+i,batch,sem,idd)
                if done:
                    other.append((day, slot))
                    

    day = None
    start_slot = None
    lab_assigned=None
    if evening_labp1:
        day, start_slot = random.choice(evening_labp1)
        lab_assigned=course.lab_id1
    elif morning_labp1:
        day,start_slot = random.choice(morning_labp1)
        lab_assigned=course.lab_id1
    elif evening_labp2:
        day,start_slot = random.choice(evening_labp2)
        lab_assigned=course.lab_id2
    elif morning_labp2:
        day,start_slot = random.choice(morning_labp2)
        lab_assigned=course.lab_id2
    elif evening_labp3:
        day,start_slot = random.choice(evening_labp3)
        lab_assigned=course.lab_id3
    elif morning_labp3:
        day,start_slot = random.choice(morning_labp3)
        lab_assigned=course.lab_id3
    elif other:
        day,start_slot=random.choice(other)
        labsss=find_available_lab(day,start_slot,-1, batch,sem,course.lab_hour)
        lab_assigned=labsss.id
    else:
        print(f"No available slots for course {course.name}")
        return
    try:
        # print(course.batch_id)
        # print(course.id)
        # print(course.lab_professor_id)
        # print(lab_assigned)
        # print(day)
        # print(slot)
        # print(sem)
        for offset in range(course.lab_hour):
            new_schedule = Schedule(
                batch_id=course.batch_id,
                course_id=course.id,
                professor_id = course.lab_professor_id if course.lab_professor_id else -1,
                lab_id=lab_assigned,
                day=day,
                slot=start_slot + offset,
                classroom_id= None,
                semester=sem,
                divide_id=idd
            )
            db.session.add(new_schedule)
        print("DONNNN")
        db.session.commit()
        print("SUCCESS")
        flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
    except Exception as e:
        db.session.rollback()
        import traceback
        traceback.print_exc()  # full error trace
        print("Error:", e)

def assign_Morning_lab(course,idd,sem=False):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    morning_labp1 = []
    evening_labp1 = []
    morning_labp2 = []
    evening_labp2 = []
    morning_labp3 = []
    evening_labp3 = []
    other=[]

    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=0:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority1(course, day, slot+i,batch,sem,idd)
                if done:
                    morning_labp1.append((day, slot))
                    
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(6,9):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            done=True
            for i in range(course.lab_hour):
                if slot+i==5 or slot+i==0 or slot+i==10:
                    done=False
                done=done and is_slot_available_lab_priority1(course, day, slot+i,batch,sem,idd)
            if done:
                evening_labp1.append((day, slot))
                
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=0:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority2(course, day, slot+i,batch,sem,idd)
                if done:
                    morning_labp2.append((day, slot))
                    
    
    for day in range(5):
            if day == course.avoid_day:
                continue 
            for slot in range(6,9):
                if slot==0 or slot==2 or slot==5 or slot==7:
                    continue
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority2(course, day, slot+i,batch,sem,idd)
                if done:
                    evening_labp2.append((day, slot))
                    
    
    for day in range(5):
        if day == course.avoid_day:
            continue 
        for slot in range(4):
            if slot==0 or slot==2 or slot==5 or slot==7:
                continue
            if slot != 4 and slot!=0:
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority3(course, day, slot+i,batch,sem,idd)
                if done:
                    morning_labp3.append((day, slot))
                    
    
    for day in range(5):
            if day == course.avoid_day:
                continue 
            for slot in range(6,9):
                if slot==0 or slot==2 or slot==5 or slot==7:
                    continue
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab_priority3(course, day, slot+i,batch,sem,idd)
                if done:
                    evening_labp3.append((day, slot))
                    
    
    for day in range(5): 
        for slot in range(9): 
            if slot != 4 and slot!=5:
                if slot==0 or slot==2 or slot==5 or slot==7:
                    continue
                done=True
                for i in range(course.lab_hour):
                    if slot+i==5 or slot+i==0 or slot+i==10:
                        done=False
                    done=done and is_slot_available_lab(course, day, slot+i,batch,sem,idd)
                if done:
                    other.append((day, slot))
                    

    day = None
    start_slot = None
    lab_assigned=None
    if morning_labp1:
        day, start_slot = random.choice(morning_labp1)
        lab_assigned=course.lab_id1
    elif evening_labp1:
        day,start_slot = random.choice(evening_labp1)
        lab_assigned=course.lab_id1
    elif morning_labp2:
        day,start_slot = random.choice(morning_labp2)
        lab_assigned=course.lab_id2
    elif evening_labp2:
        day,start_slot = random.choice(evening_labp2)
        lab_assigned=course.lab_id2
    elif morning_labp3:
        day,start_slot = random.choice(morning_labp3)
        lab_assigned=course.lab_id3
    elif evening_labp3:
        day,start_slot = random.choice(evening_labp3)
        lab_assigned=course.lab_id3
    elif other:
        day,start_slot=random.choice(other)
        lab_assigned=find_available_lab(day,start_slot,-1, batch,sem,course.lab_hour)
    else:
        print(f"No available slots for course {course.name}")
        return  # Skip this iteration if no slots available
    for offset in range(course.lab_hour):
        new_schedule = Schedule(
            batch_id=course.batch_id,
            course_id=course.id,
            professor_id = course.lab_professor_id if course.lab_professor_id else -1,
            lab_id=lab_assigned,
            day=day,
            slot=start_slot + offset,
            classroom_id= None,
            semester=sem,
            divide_id=idd
        )
        db.session.add(new_schedule)
    try:
        db.session.commit()
        flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
    except:
        db.session.rollback()
        flash('Error scheduling priority course (2-hour consecutive)', 'error')

# access as course["lab"]
def assign_priority_morning_courses(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()
    priority_priority1_slots=[]
    priority_priority2_slots=[]
    priority_priority3_slots=[]
    priority_other_slots=[]

    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()

    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5):  # Check up to slot 7 for consecutive slots
            if slot != 4 and slot !=5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_1,batch,sem):
                    priority_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_2,batch,sem):
                    priority_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_3,batch,sem):
                    priority_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem) and is_slot_available(course, day, slot+1,-1,batch,sem):
                    priority_other_slots.append((day,slot))

    days_done=[]

    while course["lecture_hour"]>1 and ( priority_priority1_slots or priority_priority2_slots or priority_priority3_slots or priority_other_slots):
            day = None
            start_slot = None
            
            if priority_priority1_slots:
                day, start_slot = random.choice(priority_priority1_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_priority2_slots:
                day,start_slot = random.choice(priority_priority2_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_priority3_slots:
                day,start_slot = random.choice(priority_priority3_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_other_slots:
                day,start_slot = random.choice(priority_other_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            else:
                print(f"No available slots for course {course['name']}")
                continue  # Skip this iteration if no slots available
            classroom1 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
            classroom2 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
            classroom3 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
            last=find_available_classroom(day,slot,batch,sem)

            if day in days_done:
                continue

            days_done.append(day)
            decided_classroom=None
            if classroom1:
                decided_classroom=classroom1
            elif classroom2:
                decided_classroom=classroom2
            elif classroom3:
                decided_classroom=classroom3
            else:
                decided_classroom=last
            


            course["lecture_hour"]-=2
            for offset in range(2):
                new_schedule = Schedule(
                    batch_id=course["batch_id"],
                    course_id=course["id"],
                    professor_id=course["professor_id"],
                    day=day,
                    slot=start_slot + offset,
                    classroom_id= decided_classroom.id,
                    semester=sem
                )
                db.session.add(new_schedule)
            try:
                db.session.commit()
                flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
            except:
                db.session.rollback()
                flash('Error scheduling priority course (2-hour consecutive)', 'error')

    morning_priority1_slots=[]
    morning_priority2_slots=[]
    morning_priority3_slots=[]
    morning_other_slots=[]

    for day in range(5):
            if day == course["avoid_day"] or day in days_done:
                continue 
            for slot in range(5,10):  # Check up to slot 7 for consecutive slots
                if slot !=5 and not (not sem and slot==0):
                    if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                        morning_priority1_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                        morning_priority2_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                        morning_priority3_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,-1,batch,sem):
                        morning_other_slots.append((day,slot))
    
    evening_priority1_slots=[]
    evening_priority2_slots=[]
    evening_priority3_slots=[]
    evening_other_slots=[]

    for day in range(5):
            if day == course["avoid_day"] or day in days_done:
                continue 
            for slot in range(5):  # Check up to slot 7 for consecutive slots
                if slot !=5 and not (not sem and slot==0):
                    if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                        evening_priority1_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                        evening_priority2_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                        evening_priority3_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,-1,batch,sem):
                        evening_other_slots.append((day,slot))

    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot !=5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_other_slots or last_priority1_slots or last_priority2_slots or last_priority3_slots or morning_priority1_slots or morning_priority2_slots or morning_priority3_slots or morning_other_slots or evening_other_slots or evening_priority1_slots or evening_priority2_slots or evening_priority3_slots):
        day = None
        start_slot = None
        
        if evening_priority1_slots:
            day, start_slot = random.choice(evening_priority1_slots)
            evening_priority1_slots.remove((day, start_slot))
        elif evening_priority2_slots:
            day,start_slot = random.choice(evening_priority2_slots)
            evening_priority2_slots.remove((day, start_slot))
        elif evening_priority3_slots:
            day,start_slot = random.choice(evening_priority3_slots)
            evening_priority3_slots.remove((day, start_slot))
        elif evening_other_slots:
            day,start_slot = random.choice(evening_other_slots)
            evening_other_slots.remove((day, start_slot))
        elif morning_priority1_slots:
            day,start_slot=random.choice(morning_priority1_slots)
            morning_priority1_slots.remove((day,start_slot))
        elif morning_priority2_slots:
            day,start_slot = random.choice(morning_priority2_slots)
            morning_priority2_slots.remove((day, start_slot))
        elif morning_priority3_slots:
            day,start_slot = random.choice(morning_priority3_slots)
            morning_priority3_slots.remove((day, start_slot))
        elif morning_other_slots:
            day,start_slot = random.choice(morning_other_slots)
            morning_other_slots.remove((day, start_slot))
        elif last_priority1_slots:
            day,start_slot=random.choice(last_priority1_slots)
            last_priority1_slots.remove((day,start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
        decided_classroom=None
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        if day in days_done:
            continue
        days_done.append(day)
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        elif last:
            decided_classroom=last
        else:
            continue
        
        course["lecture_hour"]-=1
        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_priority_evening_courses(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()
    priority_priority1_slots=[]
    priority_priority2_slots=[]
    priority_priority3_slots=[]
    priority_other_slots=[]
    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()

    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5,9):
            if slot != 4 and slot !=5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_1,batch,sem):
                    priority_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_2,batch,sem):
                    priority_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_3,batch,sem):
                    priority_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem) and is_slot_available(course, day, slot+1,-1,batch,sem):
                    priority_other_slots.append((day,slot))

    days_done=[]

    while course["lecture_hour"]>1 and ( priority_priority1_slots or priority_priority2_slots or priority_priority3_slots or priority_other_slots):
            day = None
            start_slot = None
            if priority_priority1_slots:
                day, start_slot = random.choice(priority_priority1_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_priority2_slots:
                day,start_slot = random.choice(priority_priority2_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_priority3_slots:
                day,start_slot = random.choice(priority_priority3_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            elif priority_other_slots:
                day,start_slot = random.choice(priority_other_slots)
                priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
                priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
                priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
                priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
            else:
                print(f"No available slots for course {course['name']}")
                continue  # Skip this iteration if no slots available
            classroom1 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
            classroom2 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
            classroom3 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
            last=find_available_classroom(day,slot,batch,sem)

            if day in days_done:
                continue
            days_done.append(day)
            decided_classroom=None
            if classroom1:
                decided_classroom=classroom1
            elif classroom2:
                decided_classroom=classroom2
            elif classroom3:
                decided_classroom=classroom3
            else:
                decided_classroom=last
            
            course["lecture_hour"]-=2
            for offset in range(2):
                new_schedule = Schedule(
                    batch_id=course["batch_id"],
                    course_id=course["id"],
                    professor_id=course["professor_id"],
                    day=day,
                    slot=start_slot + offset,
                    classroom_id= decided_classroom.id,
                    semester=sem
                )
                db.session.add(new_schedule)
            try:
                db.session.commit()
                flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
            except:
                db.session.rollback()
                flash('Error scheduling priority course (2-hour consecutive)', 'error')

    morning_priority1_slots=[]
    morning_priority2_slots=[]
    morning_priority3_slots=[]
    morning_other_slots=[]

    for day in range(5):
            if day == course["avoid_day"] or day in days_done:
                continue 
            for slot in range(5):  # Check up to slot 7 for consecutive slots
                if slot !=5 and not (not sem and slot==0):
                    if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                        morning_priority1_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                        morning_priority2_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                        morning_priority3_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,-1,batch,sem):
                        morning_other_slots.append((day,slot))
    
    evening_priority1_slots=[]
    evening_priority2_slots=[]
    evening_priority3_slots=[]
    evening_other_slots=[]

    for day in range(5):
            if day == course["avoid_day"] or day in days_done:
                continue 
            for slot in range(5,10):  # Check up to slot 7 for consecutive slots
                if slot !=5 and not (not sem and slot==0):
                    if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                        evening_priority1_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                        evening_priority2_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                        evening_priority3_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,-1,batch,sem):
                        evening_other_slots.append((day,slot))

    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot !=5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_other_slots or last_priority1_slots or last_priority2_slots or last_priority3_slots or morning_priority1_slots or morning_priority2_slots or morning_priority3_slots or morning_other_slots or evening_other_slots or evening_priority1_slots or evening_priority2_slots or evening_priority3_slots):
        day = None
        start_slot = None
        if evening_priority1_slots:
            day, start_slot = random.choice(evening_priority1_slots)
            evening_priority1_slots.remove((day, start_slot))
        elif evening_priority2_slots:
            day,start_slot = random.choice(evening_priority2_slots)
            evening_priority2_slots.remove((day, start_slot))
        elif evening_priority3_slots:
            day,start_slot = random.choice(evening_priority3_slots)
            evening_priority3_slots.remove((day, start_slot))
        elif evening_other_slots:
            day,start_slot = random.choice(evening_other_slots)
            evening_other_slots.remove((day, start_slot))
        elif morning_priority1_slots:
            day,start_slot=random.choice(morning_priority1_slots)
            morning_priority1_slots.remove((day,start_slot))
        elif morning_priority2_slots:
            day,start_slot = random.choice(morning_priority2_slots)
            morning_priority2_slots.remove((day, start_slot))
        elif morning_priority3_slots:
            day,start_slot = random.choice(morning_priority3_slots)
            morning_priority3_slots.remove((day, start_slot))
        elif morning_other_slots:
            day,start_slot = random.choice(morning_other_slots)
            morning_other_slots.remove((day, start_slot))
        elif last_priority1_slots:
            day,start_slot=random.choice(last_priority1_slots)
            last_priority1_slots.remove((day,start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available

        if day in days_done:
            continue
        days_done.append(day)
        decided_classroom=None
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        elif last:
            decided_classroom=last
        else:
            continue
        
        course["lecture_hour"]-=1
        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot ,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    # last
    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')


    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_only_priority(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()

    priority_priority1_slots=[]
    priority_priority2_slots=[]
    priority_priority3_slots=[]
    priority_other_slots=[]
    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()


    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(9):  # Check up to slot 7 for consecutive slots
            if slot != 5 and slot !=4 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_1,batch,sem):
                    priority_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_2,batch,sem):
                    priority_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem) and is_slot_available(course, day, slot+1,specific_professor.priority_classroom_3,batch,sem):
                    priority_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem) and is_slot_available(course, day, slot+1,-1,batch,sem):
                    priority_other_slots.append((day,slot))

    days_done=[]

    while course["lecture_hour"]>1 and (priority_priority1_slots or priority_priority2_slots or priority_priority3_slots or priority_other_slots):
        day = None
        start_slot = None
        if priority_priority1_slots:
            day, start_slot = random.choice(priority_priority1_slots)
            priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
            priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
            priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
            priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
        elif priority_priority2_slots:
            day,start_slot = random.choice(priority_priority2_slots)
            priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
            priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
            priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
            priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
        elif priority_priority3_slots:
            day,start_slot = random.choice(priority_priority3_slots)
            priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
            priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
            priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
            priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
        elif priority_other_slots:
            day,start_slot = random.choice(priority_other_slots)
            priority_priority1_slots = [slot for slot in priority_priority1_slots if slot[0] != day]
            priority_priority2_slots = [slot for slot in priority_priority2_slots if slot[0] != day]
            priority_priority3_slots = [slot for slot in priority_priority3_slots if slot[0] != day]
            priority_other_slots = [slot for slot in priority_other_slots if slot[0] != day]
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available

        if day in days_done:
            continue
        days_done.append(day)

        decided_classroom=None
        classroom1 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom(day,slot,batch,sem)
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        
        course["lecture_hour"]-=2
        for offset in range(2):
            new_schedule = Schedule(
                batch_id=course["batch_id"],
                course_id=course["id"],
                professor_id=course["professor_id"],
                day=day,
                slot=start_slot + offset,
                classroom_id= decided_classroom.id,
                semester=sem
            )
            db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    priority1_slots=[]
    priority2_slots=[]
    priority3_slots=[]
    other_slots=[]

    for day in range(5):
            if day == course["avoid_day"] or day in days_done:
                continue 
            for slot in range(10):  # Check up to slot 7 for consecutive slots
                if slot !=5 and not (not sem and slot==0):
                    if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                        priority1_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                        priority2_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                        priority3_slots.append((day, slot))
                    elif is_slot_available(course, day, slot,-1,batch,sem):
                        other_slots.append((day,slot))
        

    while course["lecture_hour"]>0 and (priority1_slots or priority2_slots or priority3_slots or other_slots):
        day = None
        start_slot = None
        if priority1_slots:
            day, start_slot = random.choice(priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif priority2_slots:
            day,start_slot = random.choice(priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif priority3_slots:
            day,start_slot = random.choice(priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif other_slots:
            day,start_slot = random.choice(other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        if day in days_done:
            continue
        days_done.append(day)
        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_morning_only(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    priority1_slots=[]
    priority2_slots=[]
    priority3_slots=[]
    other_slots=[]
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()
    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()

    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5):  # Check up to slot 7 for consecutive slots
            if slot != 5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    other_slots.append((day,slot))

    evening_priority1_slots=[]
    evening_priority2_slots=[]
    evening_priority3_slots=[]
    evening_other_slots=[]

    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5,10):  # Check up to slot 7 for consecutive slots
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    evening_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    evening_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    evening_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    evening_other_slots.append((day,slot))

    days_done=[]
    while course["lecture_hour"]>0 and (priority1_slots or priority2_slots or priority3_slots or evening_priority1_slots or evening_priority2_slots or evening_priority3_slots or other_slots or evening_other_slots):
        day = None
        start_slot = None
        if priority1_slots:
            day, start_slot = random.choice(priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif priority2_slots:
            day,start_slot = random.choice(priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif priority3_slots:
            day,start_slot = random.choice(priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif other_slots:
            day,start_slot = random.choice(other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif evening_priority1_slots:
            day,start_slot = random.choice(evening_priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif evening_priority2_slots:
            day,start_slot = random.choice(evening_priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif evening_priority3_slots:
            day,start_slot = random.choice(evening_priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        elif evening_other_slots:
            day,start_slot = random.choice(evening_other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            evening_priority1_slots = [slot for slot in evening_priority1_slots if slot[0] != day]
            evening_priority2_slots = [slot for slot in evening_priority2_slots if slot[0] != day]
            evening_priority3_slots = [slot for slot in evening_priority3_slots if slot[0] != day]
            evening_other_slots = [slot for slot in evening_other_slots if slot[0] != day]
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)
                
        if day in days_done:
            continue
        days_done.append(day)
        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')
    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')


    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_evening_only(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    priority1_slots=[]
    priority2_slots=[]
    priority3_slots=[]
    other_slots=[]
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()
    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()
    days_done=[]
    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5,10):  # Check up to slot 7 for consecutive slots
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    other_slots.append((day,slot))

    morning_priority1_slots=[]
    morning_priority2_slots=[]
    morning_priority3_slots=[]
    morning_other_slots=[]

    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(5):  # Check up to slot 7 for consecutive slots
            if slot != 5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    morning_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    morning_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    morning_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    morning_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (priority1_slots or priority2_slots or priority3_slots or morning_priority1_slots or morning_priority2_slots or morning_priority3_slots or other_slots or morning_other_slots):
        day = None
        start_slot = None
        if priority1_slots:
            day, start_slot = random.choice(priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif priority2_slots:
            day,start_slot = random.choice(priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif priority3_slots:
            day,start_slot = random.choice(priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif other_slots:
            day,start_slot = random.choice(other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif morning_priority1_slots:
            day,start_slot = random.choice(morning_priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif morning_priority2_slots:
            day,start_slot = random.choice(morning_priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif morning_priority3_slots:
            day,start_slot = random.choice(morning_priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        elif morning_other_slots:
            day,start_slot = random.choice(morning_other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
            morning_priority1_slots = [slot for slot in morning_priority1_slots if slot[0] != day]
            morning_priority2_slots = [slot for slot in morning_priority2_slots if slot[0] != day]
            morning_priority3_slots = [slot for slot in morning_priority3_slots if slot[0] != day]
            morning_other_slots = [slot for slot in morning_other_slots if slot[0] != day]
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        if day in days_done:
            continue
        days_done.append(day)
        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')


    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]

    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')


    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1

        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_no_priority(course,sem=False):
    batch = Batch.query.filter(Batch.id == course["batch_id"]).first()
    priority1_slots=[]
    priority2_slots=[]
    priority3_slots=[]
    other_slots=[]
    specific_professor = Professor.query.filter(Professor.id == course["professor_id"]).first()
    if specific_professor is None:
        class DummyProf:
            priority_classroom_1 = course["p1"]
            priority_classroom_2 = course["p3"]
            priority_classroom_3 = course["p3"]

        specific_professor = DummyProf()
    print(specific_professor)
    for day in range(5):
        if day == course["avoid_day"]:
            continue 
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot != 5 and not (not sem and slot==0):
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    other_slots.append((day,slot))

    days_done=[]

    while course["lecture_hour"]>0 and (priority1_slots or priority2_slots or priority3_slots or other_slots):
        day = None
        start_slot = None
        if priority1_slots:
            day, start_slot = random.choice(priority1_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif priority2_slots:
            day,start_slot = random.choice(priority2_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif priority3_slots:
            day,start_slot = random.choice(priority3_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        elif other_slots:
            day,start_slot = random.choice(other_slots)
            priority1_slots = [slot for slot in priority1_slots if slot[0] != day]
            priority2_slots = [slot for slot in priority2_slots if slot[0] != day]
            priority3_slots = [slot for slot in priority3_slots if slot[0] != day]
            other_slots = [slot for slot in other_slots if slot[0] != day]
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
        
        if day in days_done:
            continue
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1
        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

    last_priority1_slots=[]
    last_priority2_slots=[]
    last_priority3_slots=[]
    last_other_slots=[]
    for day in range(5):
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1
        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')
    

    for day in range(6):
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot != 5:
                if is_slot_available(course, day, slot,specific_professor.priority_classroom_1,batch,sem):
                    last_priority1_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_2,batch,sem):
                    last_priority2_slots.append((day, slot))
                elif is_slot_available(course, day, slot,specific_professor.priority_classroom_3,batch,sem):
                    last_priority3_slots.append((day, slot))
                elif is_slot_available(course, day, slot,-1,batch,sem):
                    last_other_slots.append((day,slot))

    while course["lecture_hour"]>0 and (last_priority1_slots or last_priority2_slots or last_priority3_slots or last_other_slots):
        day = None
        start_slot = None
        if last_priority1_slots:
            day, start_slot = random.choice(last_priority1_slots)
            last_priority1_slots.remove((day, start_slot))
        elif last_priority2_slots:
            day,start_slot = random.choice(last_priority2_slots)
            last_priority2_slots.remove((day, start_slot))
        elif last_priority3_slots:
            day,start_slot = random.choice(last_priority3_slots)
            last_priority3_slots.remove((day, start_slot))
        elif last_other_slots:
            day,start_slot = random.choice(last_other_slots)
            last_other_slots.remove((day, start_slot))
        else:
            print(f"No available slots for course {course['name']}")
            continue  # Skip this iteration if no slots available
                
        classroom1 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_1,batch,sem)
        classroom2 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_2,batch,sem)
        classroom3 = find_available_classroom_with_priorityroom_onehour(day, start_slot,specific_professor.priority_classroom_3,batch,sem)
        last=find_available_classroom_onehour(day,slot,batch,sem)

        decided_classroom=None
        if classroom1:
            decided_classroom=classroom1
        elif classroom2:
            decided_classroom=classroom2
        elif classroom3:
            decided_classroom=classroom3
        else:
            decided_classroom=last
        course["lecture_hour"]-=1
        new_schedule = Schedule(
            batch_id=course["batch_id"],
            course_id=course["id"],
            professor_id=course["professor_id"],
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def assign_tutorial(course,idd,sem=False):
    batch = Batch.query.filter(Batch.id == course.batch_id).first()
    other_slots=[]
    for day in range(5):
        for slot in range(10):  # Check up to slot 7 for consecutive slots
            if slot != 5 and is_tutorial_slot_available(course,day,slot,batch,idd,sem):
                other_slots.append((day,slot))

    for j in range(course.tutorial_hour):
        day=None
        start_slot=None
        day,start_slot = random.choice(other_slots)
        while(not is_tutorial_slot_available(course,day,start_slot,batch,idd,sem)):
            other_slots.remove((day, start_slot))
            day,start_slot = random.choice(other_slots)

        other_slots.remove((day, start_slot))
        last=find_available_classroom_onehour(day,start_slot,batch,sem)
        decided_classroom=last
        new_schedule = Schedule(
            batch_id=course.batch_id,
            course_id=course.id,
            professor_id=-1,
            day=day,
            slot=start_slot,
            classroom_id= decided_classroom.id,
            semester=sem,
            tutorial=course.tutorial,
            divide_id=idd
        )
        db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except:
            db.session.rollback()
            flash('Error scheduling priority course (2-hour consecutive)', 'error')

def add_this_in_schedule(course,sem):
    courses=[course]
    Schedule.query.filter_by(course_id=course.id).delete()
    db.session.commit()

    professors=Professor.query.all()
    course_data = []
    for c in courses:
        course_data.append({
            "id": c.id,
            "name": c.name,
            "lab_hour": c.lab_hour,
            "lecture_hour": c.lecture_hour,
            "tutorial_hour": c.tutorial_hour,
            "is_lab": c.is_lab,
            "priority":c.priority,
            "priority_morning":c.priority_morning,
            "priority_evening":c.priority_evening,
            "avoid_day":c.avoid_day,
            "professor_id":c.professor_id,
            "lab_professor_id":c.lab_professor_id,
            "lab_id1":c.lab_id1,
            "lab_id2":c.lab_id2,
            "lab_id3":c.lab_id3,
            "batch_id":c.batch_id,
            "tutorial":c.tutorial,
            "divide":c.divide,
            "p1":c.priority_classroom_1,
            "p3":c.priority_classroom_2,
            "p3":c.priority_classroom_3
        })
    
    priority_morning_courses = [c for c in course_data if c["priority"] and c["priority_morning"]]

    priority_evening_courses = [c for c in course_data if c["priority"] and c["priority_evening"]]

    only_priority=[c for c in course_data if c["priority"] and ((not c["priority_evening"] and not c["priority_morning"]) or (c["priority_evening"] and c["priority_morning"]))]

    morning_only = [c for c in course_data if not c["priority"] and c["priority_morning"]]

    evening_only = [c for c in course_data if not c["priority"] and c["priority_evening"]]

    print("HELLO")
    no_priority=[c for c in course_data if not c["priority"] and ((not c["priority_evening"] and not c["priority_morning"]) or (c["priority_evening"] and c["priority_morning"]))]
    print("DONN")
    if request.method == 'POST':

        if course.divide:
            for i in range(2):
                if course.is_lab and course.lab_priority_evening:
                    assign_Evening_lab(course,i+1,sem)
                elif course.is_lab and course.lab_priority_morning:
                    assign_Morning_lab(course,i+1,sem)
                elif course.is_lab:
                    assign_Lab(course,i+1,sem)
        else:
            if course.is_lab and course.lab_priority_evening:
                assign_Evening_lab(course,0,sem)
            elif course.is_lab and course.lab_priority_morning:
                assign_Morning_lab(course,0,sem)
            elif course.is_lab:
                assign_Lab(course,0,sem)

        for cours in priority_morning_courses:
            assign_priority_morning_courses(cours,sem)
            
        for cours in priority_evening_courses:
            assign_priority_evening_courses(cours,sem)
            
        for cours in only_priority:
            assign_only_priority(cours,sem)
            
        for cours in morning_only:
            assign_morning_only(cours,sem)
            
        for cours in evening_only:
            assign_evening_only(cours,sem)

        for cours in no_priority:
            assign_no_priority(cours,sem)
        
        if course.divide and course.tutorial:
            for i in range(2):
                assign_tutorial(course,i+1,sem)
        elif course.tutorial:
            assign_tutorial(course,0,sem)

def make_combined_schedule(sc):
    sem=sc.is_odd
    professor=Professor.query.filter_by(id=sc.professor_id).first()
    lecture=sc.lecture
    batch_ids = [item['batch_id'] for item in sc.batch_course_list]
    capacity=sc.capacity
    slots = []
    for day in range(5):
        for slot in range(10):
            if slot != 5:
                if is_combined_available(professor, batch_ids, day, slot, sem):
                    slots.append((day, slot))

    while lecture>0 and slots:
        day = None
        start_slot = None
        day, start_slot = random.choice(slots)
        slots.remove((day, start_slot))
        classroom=find_combined_classroom(day,start_slot,sem,capacity)
        if classroom is None:
            continue
        lecture-=1

        for idd in batch_ids:
            new_schedule = Schedule(
                batch_id=idd,
                course_id=None,
                professor_id=professor.id,
                day=day,
                slot=start_slot,
                classroom_id= classroom.id,
                semester=sem,
                combined_course_id=sc.id
            )
            db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except Exception as e:
            db.session.rollback()
            print("ERROR:", e)
            flash(f"Error scheduling course: {e}", "error")

    for day in range(6):
        for slot in range(10):
            if slot != 5:
                if is_combined_available(professor, batch_ids, day, slot, sem):
                    slots.append((day, slot))

    while lecture>0 and slots:
        day = None
        start_slot = None
        day, start_slot = random.choice(slots)
        slots.remove((day, start_slot))
        classroom=find_combined_classroom(day,start_slot,sem,capacity)
        if classroom is None:
            print("SJDN")
            continue
        lecture-=1

        for idd in batch_ids:
            print("HELLO")
            new_schedule = Schedule(
                batch_id=idd,
                course_id=None,
                professor_id=professor.id,
                day=day,
                slot=start_slot,
                classroom_id= classroom.id,
                semester=sem,
                combined_course_id=sc.id
            )
            db.session.add(new_schedule)
        try:
            db.session.commit()
            flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
        except Exception as e:
            db.session.rollback()
            print("ERROR:", e)
            flash(f"Error scheduling course: {e}", "error")

def lab_available(lab,prof,batches,day,slot):
    professor=Schedule.query.filter_by(day=day,slot=slot,professor_id=prof).first()
    labs=Schedule.query.filter_by(day=day,slot=slot,lab_id=lab).first()
    if professor or labs:
        return False
    for batch in batches:
        done=Schedule.query.filter_by(day=day,slot=slot,batch_id=batch.id).first()
        if done:
            return False
    return True

def assign_lab(model):
    if int(model.lab_hour)==0:
        return
    semester=model.semester
    names=model.names.split(',')
    department_id=model.department_id
    lab_ids=model.lab_id.split(',')
    lab_prof_ids=model.lab_professor_id.split(',')
    lab_hour=model.lab_hour
    semester=int(semester)
    department_id=int(department_id)
    lab_hour=int(lab_hour)
    sem=semester%2
    batches=Batch.query.filter_by(department_id=department_id,odd_sem=sem).all()
    pp=[]
    for p in lab_ids:
        pp.append(int(p))
    lab_ids=pp
    pp=[]
    for p in lab_prof_ids:
        pp.append(int(p))
    lab_prof_ids=pp

    slots=[]
    for day in range(5):
        for slot in range(10):
            if slot==5:
                continue
            ans=True
            for lab,prof in lab_ids,lab_prof_ids:
                for i in range(lab_hour):
                    if slot+i==5 or slot+i==10 or not lab_available(lab,prof,batches,day,slot+i):
                        ans=False
            if ans:
                slots.append((day,slot))

    day,start_slot=random.choice(slots)

    for batch in batches:
        for i in range(len(lab_ids)):
            for j in range(lab_hour):
                new_schedule = Schedule(
                    name=names[i],
                    batch_id=batch.id,
                    elective_course_id=model.id,
                    professor_id=lab_prof_ids[i],
                    day=day,
                    slot=start_slot+j,
                    lab_id= lab_ids[i],
                    semester=sem,
                    elective_id=i+1
                )
                db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling priority course (2-hour consecutive)', 'error')

def class_available(prof,batches,day,slot,count):
    ids = [1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,43,44,45]
    sch=Schedule.query.filter_by(professor_id=prof,day=day,slot=slot).first()
    if sch:
        return False
    classrooms = Classroom.query.filter(Classroom.id.in_(ids)).all()

    for batch in batches:
        done=Schedule.query.filter_by(day=day,slot=slot,batch_id=batch.id).first()
        if done:
            return False
    cnt=0
    for classroom in classrooms:
        done=Schedule.query.filter_by(day=day,slot=slot,classroom_id=classroom.id).first()
        if not done:
            cnt+=1

    return cnt>=count

def find_classroom(batches,day,slot,count):
    ids = [1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,43,44,45]

    classrooms = Classroom.query.filter(Classroom.id.in_(ids)).all()

    clas=[]
    for classroom in classrooms:
        done=Schedule.query.filter_by(day=day,slot=slot,classroom_id=classroom.id).first()
        if not done:
            clas.append(classroom)

    return clas

def assign_lecture(model):
    if int(model.lecture_hour)==0:
        return
    print("hello")
    semester=model.semester
    names=model.names.split(',')
    department_id=model.department_id
    prof_ids=model.professor_id.split(',')
    lab_ids=model.lab_id.split(',')
    lab_prof_ids=model.lab_professor_id.split(',')
    lecture_hour=model.lecture_hour
    lab_hour=model.lab_hour
    tutorial_hour=model.tutorial_hour
    semester=int(semester)
    department_id=int(department_id)
    lecture_hour=int(lecture_hour)
    lab_hour=int(lab_hour)
    tutorial_hour=int(tutorial_hour)
    sem=semester%2
    batches=Batch.query.filter_by(department_id=department_id,odd_sem=sem).all()
    pp=[]
    for p in prof_ids:
        pp.append(int(p))
    prof_ids=pp
    pp=[]
    for p in lab_ids:
        pp.append(int(p))
    lab_ids=pp
    pp=[]
    for p in lab_prof_ids:
        pp.append(int(p))
    lab_prof_ids=pp

    slots=[]
    for day in range(5):
        for slot in range(10):
            if slot==5:
                continue
            ans=True
            for prof in prof_ids:
                if not class_available(prof,batches,day,slot,len(names)):
                    ans=False
            if ans:
                slots.append((day,slot))
    hour=int(model.lecture_hour)
    print(type(hour))
    while hour>0:
        day,start_slot=random.choice(slots)
        hour-=1
        slots.remove((day,start_slot))
        classroom=find_classroom(batches,day,start_slot,len(lab_ids))
        for batch in batches:
            for name, proff, i in zip(names, prof_ids, range(len(names))):
                print(batch.id)
                print(model.id)
                print(proff)
                print(day)
                print(start_slot)
                print(classroom[i])
                print(sem)
                print(i+1)
                new_schedule = Schedule(
                    name=name,
                    batch_id=batch.id,
                    elective_course_id=model.id,
                    professor_id=proff,
                    day=day,
                    slot=start_slot,
                    classroom_id= classroom[i].id,
                    semester=sem,
                    elective_id=i+1
                )
                db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling priority course (2-hour consecutive)', 'error')

def class_available_tutorial(batches,day,slot,count):
    ids = [1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,43,44,45]
    classrooms = Classroom.query.filter(Classroom.id.in_(ids)).all()

    for batch in batches:
        done=Schedule.query.filter_by(day=day,slot=slot,batch_id=batch.id).first()
        if done:
            return False
    cnt=0
    for classroom in classrooms:
        done=Schedule.query.filter_by(day=day,slot=slot,classroom_id=classroom.id).first()
        if not done:
            cnt+=1

    return cnt>=count

def find_classroom_tutorial(batches,day,slot,count):
    ids = [1,2,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,43,44,45]

    classrooms = Classroom.query.filter(Classroom.id.in_(ids)).all()

    clas=[]
    for classroom in classrooms:
        done=Schedule.query.filter_by(day=day,slot=slot,classroom_id=classroom.id).first()
        if not done:
            clas.append(classroom)

    return clas

def assign_tutorial(model):
    print("hello")
    semester=model.semester
    names=model.names.split(',')
    department_id=model.department_id
    prof_ids=model.professor_id.split(',')
    lab_ids=model.lab_id.split(',')
    lab_prof_ids=model.lab_professor_id.split(',')
    lecture_hour=model.lecture_hour
    lab_hour=model.lab_hour
    tutorial_hour=model.tutorial_hour
    semester=int(semester)
    department_id=int(department_id)
    lecture_hour=int(lecture_hour)
    lab_hour=int(lab_hour)
    tutorial_hour=int(tutorial_hour)
    sem=semester%2
    batches=Batch.query.filter_by(department_id=department_id,odd_sem=sem).all()
    pp=[]
    for p in prof_ids:
        pp.append(int(p))
    prof_ids=pp
    pp=[]
    for p in lab_ids:
        pp.append(int(p))
    lab_ids=pp
    pp=[]
    for p in lab_prof_ids:
        pp.append(int(p))
    lab_prof_ids=pp

    slots=[]
    for day in range(5):
        for slot in range(10):
            if slot==5:
                continue
            ans=True
            for prof in prof_ids:
                if not class_available_tutorial(batches,day,slot,len(names)):
                    ans=False
            if ans:
                slots.append((day,slot))
    hour=int(model.tutorial_hour)
    print(type(hour))
    while hour>0:
        day,start_slot=random.choice(slots)
        hour-=1
        slots.remove((day,start_slot))
        classroom=find_classroom_tutorial(batches,day,start_slot,len(lab_ids))
        for batch in batches:
            for name, proff, i in zip(names, prof_ids, range(len(names))):
                print(batch.id)
                print(model.id)
                print(proff)
                print(day)
                print(start_slot)
                print(classroom[i])
                print(sem)
                print(i+1)
                new_schedule = Schedule(
                    name=name,
                    batch_id=batch.id,
                    elective_course_id=model.id,
                    professor_id=proff,
                    day=day,
                    slot=start_slot,
                    classroom_id= classroom[i].id,
                    semester=sem,
                    elective_id=i+1,
                    tutorial=1
                )
                db.session.add(new_schedule)
                try:
                    db.session.commit()
                    flash('Priority course scheduled successfully (2-hour consecutive)', 'success')
                except:
                    db.session.rollback()
                    flash('Error scheduling priority course (2-hour consecutive)', 'error')


def add_elective_course(model):
    if int(model.lab_hour)>0:
        labs=model.lab_id.split(',')
        for lab in labs:
            if lab=="-1":
                print("ERROR")
                return
    assign_lab(model)
    assign_lecture(model)
    assign_tutorial(model)


# Routes
@app.route('/')
def index():
    departments=Department.query.all()
    batches = Batch.query.all()
    return render_template('index.html', batches=batches,departments=departments)
@app.route('/select_batches', methods=['GET'])
def select_batches():
    batches = Batch.query.all()
    return render_template('select_batches.html', batches=batches)

@app.route('/get_professors/<int:dept_id>')
def get_professors(dept_id):
    if dept_id == 0:
        professors = Professor.query.all()
    else:
        professors = Professor.query.filter_by(department_id=dept_id).all()
    
    return jsonify([{'id': p.id, 'name': p.name} for p in professors])

@app.route('/download-timetable/<int:batch_id>', methods=['POST','GET'])
def download_timetable(batch_id):
    batch = Batch.query.filter_by(id=batch_id).first()
    print(batch)
    if not batch:
        print('No batches selected')
        flash('No batches selected', 'error')
        return redirect(url_for('index'))
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        # for batch_id in selected_batch_ids:
        print("first")
        excel_path = generate_excel([int(batch_id)])
        print("second")
        if excel_path and os.path.exists(excel_path):
            zip_file.write(excel_path, os.path.basename(excel_path))
            os.remove(excel_path)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Batch_{batch.name}.zip'
    )

@app.route('/download-timetable-classroom/<int:prof_id>', methods=['POST','GET'])
def download_timetable_classroom(prof_id):
    print("CORRECT")
    if not prof_id:
        flash('No batches selected', 'error')
        return redirect(url_for('index'))
    
    batch=Classroom.query.filter_by(id=prof_id).first()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        # for batch_id in selected_batch_ids:
        excel_path = generate_excel_classroom([int(prof_id)])
        if excel_path and os.path.exists(excel_path):
            zip_file.write(excel_path, os.path.basename(excel_path))
            os.remove(excel_path)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Classroom_{batch.name}.zip'
    )

@app.route('/download-timetable-lab/<int:prof_id>', methods=['POST','GET'])
def download_timetable_lab(prof_id):
    print("CORRECT")
    if not prof_id:
        flash('No batches selected', 'error')
        return redirect(url_for('index'))
    batch=Lab.query.filter_by(id=prof_id).first()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        # for batch_id in selected_batch_ids:
        excel_path = generate_excel_lab([int(prof_id)])
        if excel_path and os.path.exists(excel_path):
            zip_file.write(excel_path, os.path.basename(excel_path))
            os.remove(excel_path)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Lab_{batch.name}.zip'
    )

@app.route('/download-timetable-professor/<int:prof_id>', methods=['POST','GET'])
def download_timetable_professor(prof_id):
    print("CORRECT")
    if not prof_id:
        flash('No batches selected', 'error')
        return redirect(url_for('index'))
    batch=Professor.query.filter_by(id=prof_id).first()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        # for batch_id in selected_batch_ids:
        excel_path = generate_excel_professor([int(prof_id)])
        if excel_path and os.path.exists(excel_path):
            zip_file.write(excel_path, os.path.basename(excel_path))
            os.remove(excel_path)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name=f'Professor_{batch.name}.zip'
    )

@app.route('/download-timetable-all-batches/<int:sem>', methods=['POST','GET'])
def download_timetable_all_batches(sem):
    batches = Batch.query.filter_by(odd_sem=bool(sem)).all()
    zip_buffer = BytesIO()
    print("HII")
    with ZipFile(zip_buffer, 'w') as zip_file:
        for batch in batches:
            # Generate Excel file for each batch
            excel_path = generate_excel_all_batches([batch.id])
            
            # If the Excel file was created, add it to the ZIP
            if excel_path and os.path.exists(excel_path):
                zip_file.write(excel_path, os.path.basename(excel_path))
                os.remove(excel_path)  # cleanup temporary file

    # Reset buffer position to start
    zip_buffer.seek(0)

    # Send ZIP file as a download
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='timetables_all_batches.zip'
    )

@app.route('/classroom_type/delete/<int:id>', methods=['POST'])
def delete_classroom_type(id):
    building = Building.query.get_or_404(id)
    classroom = Classroom.query.filter_by(building_id=id).all()
    for clas in classroom:
        db.session.delete(clas)
        db.session.commit()
    db.session.delete(building)
    db.session.commit()
    flash('building deleted successfully', 'danger')
    return redirect(url_for('manage_classrooms_type'))

@app.route('/classroom_type/edit/<int:id>', methods=['GET', 'POST'])
def edit_classroom_type(id):
    building = Building.query.get_or_404(id)
    if request.method == 'POST':
        building.name = request.form['name']
        db.session.commit()
        flash('Course updated successfully', 'success')
        return redirect(url_for('manage_classrooms_type'))
    return render_template('edit_classroom_type.html', building=building)

@app.route('/download-timetable-all-professors/<int:sem>', methods=['POST','GET'])
def download_timetable_all_professors(sem):
    batches=Professor.query.all()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for batch in batches:
            # Generate Excel file for each batch
            excel_path = generate_excel_all_professors([batch.id],sem)
            
            # If the Excel file was created, add it to the ZIP
            if excel_path and os.path.exists(excel_path):
                zip_file.write(excel_path, os.path.basename(excel_path))
                os.remove(excel_path)  # cleanup temporary file

    # Reset buffer position to start
    zip_buffer.seek(0)

    # Send ZIP file as a download
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='timetables_all_professors.zip'
    )

@app.route('/download-timetable-all-classrooms/<int:sem>', methods=['POST','GET'])
def download_timetable_all_classrooms(sem):
    batches=Classroom.query.all()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for batch in batches:
            # Generate Excel file for each batch
            excel_path = generate_excel_all_classrooms([batch.id],sem)
            
            # If the Excel file was created, add it to the ZIP
            if excel_path and os.path.exists(excel_path):
                zip_file.write(excel_path, os.path.basename(excel_path))
                os.remove(excel_path)  # cleanup temporary file

    # Reset buffer position to start
    zip_buffer.seek(0)

    # Send ZIP file as a download
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='timetables_all_classrooms.zip'
    )

@app.route('/download-timetable-all-labs/<int:sem>', methods=['POST','GET'])
def download_timetable_all_labs(sem):
    batches=Lab.query.all()
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, 'w') as zip_file:
        for batch in batches:
            # Generate Excel file for each batch
            excel_path = generate_excel_all_labs([batch.id],sem)
            
            # If the Excel file was created, add it to the ZIP
            if excel_path and os.path.exists(excel_path):
                zip_file.write(excel_path, os.path.basename(excel_path))
                os.remove(excel_path)  # cleanup temporary file

    # Reset buffer position to start
    zip_buffer.seek(0)

    # Send ZIP file as a download
    return send_file(
        zip_buffer,
        mimetype='application/zip',
        as_attachment=True,
        download_name='timetables_all_labs.zip'
    )

@app.route('/edit_department/<int:id>', methods=['GET', 'POST'])
def edit_department(id):
    department = Department.query.get_or_404(id)

    if request.method == 'POST':
        department.name = request.form['name']
        db.session.commit()
        flash("Department updated successfully!", "success")
        return redirect(url_for('add_department'))

    return render_template('edit_department.html', department=department)

@app.route('/add-department',methods=['POST','GET'])
def add_department():
    departments=Department.query.all()
    if request.method == "POST":
        name = request.form.get("name")
        print(name)
        if name:
            new_department = Department(name=name)
            db.session.add(new_department)

            try:
                db.session.commit()
                print("Department added successfully")
                return redirect(url_for('add_department'))
            except:
                db.session.rollback()
                flash('Error adding classroom', 'error')
    return render_template('create_department.html',departments=departments)

@app.route('/delete_department/<int:id>', methods=['POST', 'GET'])
def delete_department(id):
    classroom = Department.query.get_or_404(id)  # find by ID or show 404
    try:
        db.session.delete(classroom)   # delete it
        db.session.commit()            # save changes
        flash(f'Classroom "{classroom.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting classroom: {str(e)}', 'error')
    
    return redirect(url_for('add_department'))

@app.route('/edit_batch/<int:id>/<int:department_id>', methods=['GET', 'POST'])
def edit_batch(id,department_id):
    batch = Batch.query.get_or_404(id)

    if request.method == "POST":
        batch.name = request.form['name']
        batch.capacity = request.form['capacity']
        batch.odd_sem = True if 'odd_sem' in request.form else False

        db.session.commit()
        flash("Batch updated successfully!", "success")
        return redirect(url_for('manage_department',department_id=department_id))

    return render_template('edit_batch.html', batch=batch,department_id=department_id)

@app.route('/create_batch/<int:department_id>', methods=['GET', 'POST'])
def create_batch(department_id):
    departments=Department.query.all()
    if request.method == "POST":
        if 'file' in request.files:
            file = request.files["file"]
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 3 or not row[0].strip() or not row[1].strip() or not row[2].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                            odd_sem = str(row[2]).strip().lower() in ['true', '1', 'yes']
                            department_id=department_id
                        except ValueError:
                            invalid_count += 1
                            continue
                        
                        if Batch.query.filter_by(name=name).first():
                            duplicate_count += 1
                            continue
                        
                        new_classroom = Batch(name=name, capacity=capacity,odd_sem=odd_sem,department_id=department_id)
                        db.session.add(new_classroom)
                        success_count += 1
                    
                    db.session.commit()
                    flash_message = f"Successfully added {success_count} classrooms"
                    if duplicate_count:
                        flash_message += f", skipped {duplicate_count} duplicates"
                    if invalid_count:
                        flash_message += f", ignored {invalid_count} invalid rows"
                    flash(flash_message, 'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
        else:
            name = request.form.get("name")
            capacity = request.form.get("capacity")
            semester = request.form.get("semester")
            semester = int(semester)
            odd_sem = (semester % 2) != 0

            department_id = department_id
            if name and capacity:
                new_classroom = Batch(name=name, capacity=capacity,odd_sem=odd_sem,department_id=department_id,semester=semester)
                db.session.add(new_classroom)
                try:
                    db.session.commit()
                    flash('Classroom added successfully', 'success')
                except:
                    db.session.rollback()
                    flash('Error adding classroom', 'error')
    return render_template('create_batch.html',departments=departments)

@app.route('/edit_professor/<int:id>/<int:department_id>', methods=['GET', 'POST'])
def edit_professor(id, department_id):
    professor = Professor.query.get_or_404(id)
    classrooms = Building.query.all()

    if request.method == "POST":
        professor.name = request.form['name']
        professor.priority_classroom_1 = request.form.get('professor_id1') or None
        professor.priority_classroom_2 = request.form.get('professor_id2') or None
        professor.priority_classroom_3 = request.form.get('professor_id3') or None

        db.session.commit()
        flash("Professor updated successfully!", "success")

        return redirect(url_for("manage_professors", department_id=department_id))

    return render_template(
        'edit_professor.html',
        professor=professor,
        classrooms=classrooms,
        department_id=department_id
    )

@app.route('/confirm_delete_professor/<int:id>/<int:department_id>', methods=['POST', 'GET'])
def confirm_delete_professor(id,department_id):
    professor = Professor.query.get_or_404(id)  # find by ID or show 404
    professors=Professor.query.all()
    courses = Course.query.filter(
        or_(Course.professor_id == id, Course.lab_professor_id == id)
    ).all()
    batches = Batch.query.all()
    batch_map = {b.id: b.name for b in batches}

    if request.method == "POST":
        print("HEHNJDJ")
        for course in courses:
            selected_prof_id = request.form.get(str(course.id))
            if selected_prof_id == id:
                    flash("You cannot assign this professor.", "danger")
                    return redirect(request.url)
            batch=Batch.query.filter_by(id=course.batch_id).first()
            if selected_prof_id:
                selected_prof_id = int(selected_prof_id)
                Schedule.query.filter_by(course_id=course.id).delete()
                if course.professor_id==id:
                    course.professor_id = selected_prof_id
                if course.lab_professor_id==id:
                    course.lab_professor_id=selected_prof_id
                db.session.commit()
                add_this_in_schedule(course,batch.odd_sem)
            db.session.commit()
        db.session.delete(professor)
        db.session.commit()
        flash("Professor Deleted successfully!", "success")
        return redirect(url_for("manage_professors",department_id=department_id))     # change to your route

    return render_template(
        'confirm_delete_professor.html',
        professors=professors,
        courses=courses,
        batch_map=batch_map,
        id=id
    )

@app.route('/professors/<int:department_id>',methods=['GET','POST'])
def manage_professors(department_id):
    departments=Department.query.all()
    classrooms = Building.query.all()
    if request.method == "POST":
        if 'file' in request.files:
            file = request.files["file"]
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 4 or not row[0].strip() or not row[1].strip() or not row[2].strip() or not row[3].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        id1 = row[1].strip()
                        id2= row[2].strip()
                        id3= row[3].strip()

                        
                        if Professor.query.filter(
                            Professor.name == name
                        ).first():
                            duplicate_count += 1
                            continue
                        
                        # Add new professor
                        new_professor = Professor(name=name,priority_classroom_1=id1,priority_classroom_2=id2, priority_classroom_3=id3,department_id=department_id)
                        db.session.add(new_professor)
                        success_count += 1

                    db.session.commit()
                    flash_message = f"Successfully added {success_count} professors"
                    if duplicate_count:
                        flash_message += f", skipped {duplicate_count} duplicates"
                    if invalid_count:
                        flash_message += f", ignored {invalid_count} invalid rows"
                    flash(flash_message, 'success' if success_count else 'warning')

                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
        else:
            name = request.form.get("name")
           # email = request.form.get("email")
            priority1=request.form.get("professor_id1") or None
            priority2=request.form.get("professor_id2") or None
            priority3=request.form.get("professor_id3") or None
            #flash(f'{name} + {email}')
            if name:
                new_professor = Professor(name=name,priority_classroom_1=priority1,priority_classroom_2=priority2,priority_classroom_3=priority3,department_id=department_id)
                db.session.add(new_professor)
                try:
                    db.session.commit()
                    flash('Professor added successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding professor: {str(e)}', 'error')

        return redirect(url_for("manage_professors",department_id=department_id))
    professors = Professor.query.filter_by(department_id=department_id).all()
    return render_template('professors.html', professors=professors, classrooms=classrooms,departments=departments,department_id=department_id)

@app.route('/classroom_type',methods=['GET','POST'])
def manage_classrooms_type():
    if request.method=="POST":
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 1 or not row[0].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        try:    
                            if Building.query.filter_by(name=name).first():
                                duplicate_count += 1
                                continue
                                
                            new_lab = Building(name=name)
                            db.session.add(new_lab)
                            success_count += 1
                        except ValueError:
                            invalid_count += 1
                            continue
                            
                    db.session.commit()
                    flash(f'Successfully added {success_count} labs. {duplicate_count} duplicates skipped. {invalid_count} invalid rows ignored.', 
                          'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing CSV: {str(e)}', 'error')
        else:
            name = request.form.get("name")
            if name:
                new_classroom_type = Building(name=name)
                db.session.add(new_classroom_type)
                try:
                    db.session.commit()
                    flash('Classroom Type added successfully', 'success')
                except:
                    db.session.rollback()
                    flash('Error adding classroom', 'error')

        return redirect(url_for("manage_classrooms_type"))

    classrooms = Building.query.all()
    return render_template('classrooms_type.html', classrooms=classrooms)

@app.route('/edit_classroom/<int:id>/<int:idd>', methods=['GET', 'POST'])
def edit_classroom(id,idd):
    classroom = Classroom.query.get_or_404(id)

    if request.method == 'POST':
        classroom.name = request.form['name']
        classroom.capacity = request.form['capacity']
        db.session.commit()
        flash("Classroom updated successfully!", "success")
        return redirect(url_for('manage_classrooms',idd=idd))

    return render_template('edit_classroom.html', classroom=classroom,idd=idd)

@app.route('/classrooms/<int:idd>', methods=['GET', 'POST'])
def manage_classrooms(idd):
    # print(idd)
    if request.method == "POST":
        if 'file' in request.files:
            file = request.files["file"]
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 2 or not row[0].strip() or not row[1].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                        except ValueError:
                            invalid_count += 1
                            continue
                        
                        if Classroom.query.filter_by(name=name).first():
                            duplicate_count += 1
                            continue
                        
                        new_classroom = Classroom(name=name, capacity=capacity, building_id=idd)
                        db.session.add(new_classroom)
                        success_count += 1
                    
                    db.session.commit()
                    flash_message = f"Successfully added {success_count} classrooms"
                    if duplicate_count:
                        flash_message += f", skipped {duplicate_count} duplicates"
                    if invalid_count:
                        flash_message += f", ignored {invalid_count} invalid rows"
                    flash(flash_message, 'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
        else:
            name = request.form.get("name")
            capacity = request.form.get("capacity")
            if name and capacity:
                new_classroom = Classroom(name=name, capacity=capacity,building_id=idd)
                db.session.add(new_classroom)
                try:
                    db.session.commit()
                    flash('Classroom added successfully', 'success')
                except:
                    db.session.rollback()
                    flash('Error adding classroom', 'error')

        return redirect(url_for("manage_classrooms",idd=idd))
    
    classrooms = Classroom.query.filter_by(building_id=idd).all()
    return render_template('classrooms.html', classrooms=classrooms, idd=idd)

@app.route('/delete_classroom/<int:id>', methods=['POST', 'GET'])
def delete_classroom(id):
    classroom = Classroom.query.get_or_404(id)  # find by ID or show 404
    idd=classroom.building_id
    try:
        db.session.delete(classroom)   # delete it
        db.session.commit()            # save changes
        flash(f'Classroom "{classroom.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting classroom: {str(e)}', 'error')
    
    return redirect(url_for('manage_classrooms',idd=idd))

@app.route('/delete_lab/<int:id>', methods=['POST', 'GET'])
def delete_lab(id):
    lab = Lab.query.get_or_404(id)  # find by ID or show 404
    try:
        db.session.delete(lab)   # delete it
        db.session.commit()            # save changes
        flash(f'Classroom "{lab.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error deleting classroom: {str(e)}', 'error')
    
    return redirect(url_for('manage_labs'))

@app.route('/edit_lab/<int:id>', methods=['GET', 'POST'])
def edit_lab(id):
    lab = Lab.query.get_or_404(id)

    if request.method == 'POST':
        lab.name = request.form['name']
        lab.capacity = request.form['capacity']
        db.session.commit()
        flash("Lab updated successfully!", "success")
        return redirect(url_for('manage_labs'))

    return render_template('edit_lab.html', lab=lab)

@app.route('/labs', methods=['GET', 'POST'])
def manage_labs():
    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files['file']
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 2 or not row[0].strip() or not row[1].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                            if capacity <= 0:
                                invalid_count += 1
                                continue
                                
                            if Lab.query.filter_by(name=name).first():
                                duplicate_count += 1
                                continue
                                
                            new_lab = Lab(name=name, capacity=capacity)
                            db.session.add(new_lab)
                            success_count += 1
                        except ValueError:
                            invalid_count += 1
                            continue
                            
                    db.session.commit()
                    flash(f'Successfully added {success_count} labs. {duplicate_count} duplicates skipped. {invalid_count} invalid rows ignored.', 
                          'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing CSV: {str(e)}', 'error')
        else:
            lab_name = request.form.get('lab_name')
            lab_capacity = request.form.get('lab_capacity')
            
            if lab_name and lab_capacity:
                try:
                    capacity = int(lab_capacity)
                    if capacity <= 0:
                        flash('Capacity must be positive', 'error')
                    elif Lab.query.filter_by(name=lab_name).first():
                        flash('Lab already exists', 'error')
                    else:
                        new_lab = Lab(name=lab_name, capacity=capacity)
                        db.session.add(new_lab)
                        db.session.commit()
                        flash('Lab added successfully', 'success')
                except ValueError:
                    flash('Invalid capacity value', 'error')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding lab: {str(e)}', 'error')
                    
        return redirect(url_for('manage_labs'))
    
    labs = Lab.query.order_by(Lab.name).all()
    return render_template('labs.html', labs=labs)

@app.route('/delete_batch/<int:id>', methods=['POST', 'GET'])
def delete_batch(id):
    print(id)
    batch = Batch.query.get_or_404(id) 
    print(batch)
    try:
        courses= Course.query.filter_by(batch_id=id).all()
        for course in courses:
            Schedule.query.filter_by(course_id=course.id).delete(synchronize_session=False)
            db.session.delete(course)
        db.session.delete(batch)
        db.session.commit()
        flash(f'Batch "{batch.name}" deleted successfully.', 'success')
    except Exception as e:
        db.session.rollback()            # Undo changes if error occurs
        flash(f'Error deleting batch: {str(e)}', 'error')
    
    return redirect(url_for('index'))

@app.route('/departments/<int:department_id>', methods=['GET', 'POST'])
def manage_department(department_id):
    classrooms = Building.query.all()
    batch = Batch.query.filter_by(department_id=department_id).all()
    return render_template(
        'batches.html',
        batches=batch,
        department_id=department_id,
        classrooms=classrooms
    )

@app.route('/elective_course/<int:department_id>',methods=['GET','POST'])
def elective_course(department_id):
    professors=Professor.query.all()
    department = Department.query.get_or_404(department_id)
    labs=Lab.query.all()
    electives = Elective.query.filter_by(department_id=department_id).order_by(Elective.semester, Elective.names).all()
    if request.method == 'POST':
        model = request.form.getlist('names[]')
        prof_ids = request.form.getlist('professors[]')        # list of professor ids (strings)
        lab_ids = request.form.getlist('labs[]')               # list of lab ids (strings)
        lab_prof_ids = request.form.getlist('lab_professors[]')# list of lab professor ids (strings)
        semester = request.form.get("semester")
        lectureHour = int(request.form['lectureHour'])
        labHour = int(request.form['labHour'])
        tutorialHour = int(request.form['tutorialHour'])
        model = Elective(names=",".join(model),semester=semester,department_id=department_id,
                         lecture_hour=lectureHour,lab_hour=labHour,tutorial_hour=tutorialHour,
                         professor_id=",".join(prof_ids),lab_id=",".join(lab_ids),lab_professor_id=",".join(lab_prof_ids))
        db.session.add(model)
        if int(model.lab_hour)>0:
            labs=model.lab_id.split(',')
            for lab in labs:
                if lab=="-1":
                    print("ERROR")
                    return redirect(url_for('elective_course', department_id=department_id))
        db.session.commit()
        add_elective_course(model)
    return render_template('add_elective.html',electives=electives,department_id=department_id,professors=professors,labs=labs)


@app.route('/delete_combined_course/<int:id>', methods=['POST'])
def delete_combined_course(id):
    #print(id)
    Schedule.query.filter_by(combined_course_id=id).delete()
    course = CombinedCourse.query.get_or_404(id)

    try:
        db.session.delete(course)
        db.session.commit()
        flash("Combined Course deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash("Error deleting combined course: " + str(e), "danger")

    return redirect(url_for('com'))

@app.route('/com',methods=['GET','POST'])
def com():
    batches=Batch.query.all()
    professors=Professor.query.all()
    combined = CombinedCourse.query.all()
    batch_dict = {b.id: b.name for b in batches}
    prof_dict = {p.id: p.name for p in professors}

    print(combined)
    if request.method == "POST":

        name = request.form.get("name")
        is_odd = True if request.form.get("is_odd") else False
        professor_id = request.form.get("professor_id")
        lecture = int(request.form['lecture'])
        capacity = int(request.form['capacity'])

        batch_ids = request.form.getlist("batch_id")
        course_codes = request.form.getlist("course_code")

        # Build JSON list
        batch_course_data = []
        for b, c in zip(batch_ids, course_codes):
            if b and c:
                batch_course_data.append({
                    "batch_id": int(b),
                    "course_code": c
                })

        # Validation
        if not name:
            flash("Name is required!", "danger")
            return redirect(request.url)

        if not professor_id:
            flash("Professor is required!", "danger")
            return redirect(request.url)

        if len(batch_course_data) == 0:
            flash("At least one batch must be added!", "danger")
            return redirect(request.url)

        # Save to DB
        sc = CombinedCourse(
            name=name,
            is_odd=is_odd,
            professor_id=professor_id,
            lecture=lecture,
            capacity=capacity,
            batch_course_list=batch_course_data
        )

        db.session.add(sc)
        db.session.commit()
        print("BEFORE")
        make_combined_schedule(sc)

        flash("Special course created!", "success")
        return redirect(url_for("com"))
    return render_template(
        'combined_class.html',
        batches=batches,
        professors=professors,
        combined_courses=combined,
        batch_dict=batch_dict,
        prof_dict=prof_dict
    )

@app.route('/batch/<int:batch_id>', methods=['GET', 'POST'])
def manage_batch(batch_id):
    classrooms = Building.query.all()
    professors = Professor.query.all()
    labs = Lab.query.all() 
    batch = Batch.query.get_or_404(batch_id)
    courses = Course.query.filter_by(batch_id=batch_id).all()
    sem=batch.odd_sem
    departments=Department.query.all()
    department=Department.query.filter_by(id=batch.department_id).first()
    all_combined = CombinedCourse.query.all()

    matching = [
        cc for cc in all_combined
        if any(item["batch_id"] == batch_id for item in cc.batch_course_list)
    ]


    if request.method == 'POST':
        if 'file' in request.files:
            file = request.files["file"]
            if file.filename != '':
                try:
                    stream = TextIOWrapper(file.stream)
                    csv_input = csv.reader(stream)
                    next(csv_input)
                    
                    success_count = 0
                    duplicate_count = 0
                    invalid_count = 0
                    
                    for row in csv_input:
                        if len(row) < 3 or not row[0].strip() or not row[1].strip() or not row[2].strip():
                            invalid_count += 1
                            continue
                        
                        name = row[0].strip()
                        try:
                            capacity = int(row[1])
                            idd = int(row[2])
                        except ValueError:
                            invalid_count += 1
                            continue
                        
                        if Classroom.query.filter_by(name=name).first():
                            duplicate_count += 1
                            continue
                        
                        new_classroom = Classroom(name=name, capacity=capacity, building_id=idd)
                        db.session.add(new_classroom)
                        success_count += 1
                    
                    db.session.commit()
                    flash_message = f"Successfully added {success_count} classrooms"
                    if duplicate_count:
                        flash_message += f", skipped {duplicate_count} duplicates"
                    if invalid_count:
                        flash_message += f", ignored {invalid_count} invalid rows"
                    flash(flash_message, 'success' if success_count else 'warning')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error processing file: {str(e)}', 'error')
        else:
            priority1=request.form.get("professor_id1") or None
            priority2=request.form.get("professor_id2") or None
            priority3=request.form.get("professor_id3") or None
            course_name = request.form['course_name']
            course_code = request.form['course_code'] or department.name
            lectureHour = int(request.form['lectureHour'])
            labHour = int(request.form['labHour'])
            tutorialHour = int(request.form['tutorialHour'])
            professor_id = request.form.get('professor_id')
            lab_professor_id=request.form.get('professor_id_lab')
            is_lab = 'is_lab' in request.form
            priority = 'priority' in request.form
            priority_shift='priority_day_type' in request.form
            priority_day='priority_day' in request.form
            avoid_day = request.form.get('avoid_day')
            lab_classroom_id1 = request.form.get('lab_classroom_id1') if is_lab else None
            lab_classroom_id2 = request.form.get('lab_classroom_id2') if is_lab else None
            lab_classroom_id3 = request.form.get('lab_classroom_id3') if is_lab else None
            lab_priority_evening='priority_lab_evening' in request.form
            lab_priority_day='priority_lab_day' in request.form
            divide='divide' in request.form
            tutorial='tutorial' in request.form

            if avoid_day:
                avoid_day = int(avoid_day)

            if course_name :
                new_course = Course(
                    name=course_name,
                    lecture_hour=int(lectureHour),
                    lab_hour=int(labHour),
                    tutorial_hour=int(tutorialHour),
                    professor_id=professor_id,
                    is_lab=is_lab,
                    priority=priority,
                    avoid_day=avoid_day,
                    priority_morning=priority_day,
                    batch_id=batch_id,
                    priority_evening=priority_shift,
                    lab_professor_id=lab_professor_id,
                    lab_id1=int(lab_classroom_id1) if lab_classroom_id1 else None,
                    lab_id2=int(lab_classroom_id2) if lab_classroom_id2 else None,
                    lab_id3=int(lab_classroom_id3) if lab_classroom_id3 else None,
                    lab_priority_morning=lab_priority_day,
                    lab_priority_evening=lab_priority_evening,
                    divide=divide,
                    course_code=course_code,
                    tutorial=tutorial,
                    priority_classroom_1=priority1,
                    priority_classroom_2=priority2,
                    priority_classroom_3=priority3
                )
                try:
                    db.session.add(new_course)
                    db.session.commit()
                    #time.sleep(3)
                    add_this_in_schedule(new_course,sem)
                    flash('Course added successfully', 'success')
                except Exception as e:
                    db.session.rollback()
                    flash(f'Error adding course: {str(e)}', 'error')
            else:
                flash('Please fill in all required fields', 'warning')
            courses = Course.query.filter_by(batch_id=batch_id).all()
        
    # --- Timetable logic ---
    days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday","Saturday"]
    time_slots = [
        "8:00 - 9:00", "9:00 - 10:00", "10:00 - 11:00", "11:00 - 12:00",
        "12:00 - 13:00", "13:00 - 14:00", "14:00 - 15:00", "15:00 - 16:00", "16:00 - 17:00", "17:00 - 18:00"
    ]

    timetable = [["" for _ in range(len(time_slots))] for _ in range(len(days))]

    schedules = Schedule.query.filter_by(batch_id=batch_id).all()

    for sch in schedules:
        course = Course.query.get(sch.course_id)
        if not course:
            course=CombinedCourse.query.filter_by(id=sch.combined_course_id).first()
        if not course:
            continue
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry=f"{course.name}"
        else :
            entry = f"{course.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"


        if 0 <= sch.day < len(days) and 0 <= sch.slot < len(time_slots):
            timetable[sch.day][sch.slot] += entry

    for sch in schedules:
        if sch.elective_id==0:
            continue
        print(sch.id)
        course = Elective.query.get(sch.elective_course_id)
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry+=f"{sch.name} "
        else :
            entry += f"{sch.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"
        timetable[sch.day][sch.slot] += entry
                 
    # return render_template('manage_batch.html', professors=professors, classrooms=labs,batch=batch,courses=courses)
    return render_template(
        'manage_batch.html',
        professors=professors,
        classrooms=labs,
        batch=batch,
        courses=courses,
        days=days,
        time_slots=time_slots,
        departments=departments,
        timetable=timetable,
        matching=matching,
        classroomss=classrooms
    )

@app.route('/change-batch-timetable/<int:id>',methods=['GET','POST'])
def change_batch_timetable(id):
    if request.method == 'POST':
        print("HELLO")
        db.session.query(Schedule).filter(
            Schedule.batch_id == id,
            Schedule.combined_course_id.is_(None)
        ).delete()
        db.session.commit()

        courses=Course.query.filter_by(batch_id=id).all()
        batch=Batch.query.filter_by(id=id).first()
        for course in courses:
            print("HII")
            add_this_in_schedule(course,batch.odd_sem)
    return redirect(url_for('manage_batch', batch_id=id))    

@app.route('/change-course-timetable/<int:batch_id>/<int:course_id>',methods=['GET','POST'])
def change_course_timetable(batch_id,course_id):
    if request.method == 'POST':
        print("HELLO")
        db.session.query(Schedule).filter(
            Schedule.batch_id == batch_id,
            Schedule.course_id==course_id,
            Schedule.combined_course_id.is_(None)
        ).delete()
        db.session.commit()

        courses=Course.query.filter_by(id=course_id).all()
        batch=Batch.query.filter_by(id=batch_id).first()
        for course in courses:
            add_this_in_schedule(course,batch.odd_sem)
    return redirect(url_for('manage_batch', batch_id=batch_id))    

@app.route('/edit_course/<int:id>/<int:batch_id>', methods=['GET', 'POST'])
def edit_course(id, batch_id):
    course = Course.query.get_or_404(id)

    professors = Professor.query.all()
    classrooms = Building.query.all()

    if request.method == "POST":
        course.name = request.form["course_name"]
        course.course_code = request.form["course_code"]
        db.session.commit()
        flash("Course updated successfully!", "success")
        return redirect(url_for("manage_batch", batch_id=batch_id))

    return render_template("edit_course.html",
                           course=course,
                           professors=professors,
                           classrooms=classrooms,
                           batch_id=batch_id)

@app.route('/delete_course/<int:course_id>', methods=['POST'])
def delete_course(course_id):
    course = Course.query.get_or_404(course_id)
    schedules= Schedule.query.filter_by(course_id=course_id)
    try:
        schedules = Schedule.query.filter_by(course_id=course_id)
        schedules.delete(synchronize_session=False)
        db.session.commit()
        db.session.delete(course)
        db.session.commit()
        flash(f"Course '{course.name}' deleted successfully!", "success")
    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting course: {str(e)}", "error")

    # Redirect back to the batch page that the course belonged to
    return redirect(url_for('manage_batch', batch_id=course.batch_id))
 
@app.route('/professor_timetable/<int:sem>',methods=['GET'])
def professor_timetale(sem):
    prof_data = []
    professors = Professor.query.all()
    for p in professors:
        department=Department.query.filter_by(id=p.department_id).first()
        prof_data.append(
            {
                "id":p.id,
                "name":p.name,
                "department":department.name
            }
        )
    
    return render_template('timetable_professor.html',professors=prof_data,sem=sem)

@app.route('/batch_timetable',methods=['GET'])
def batches_timetale():
    batches = Batch.query.filter_by(odd_sem=True).all()
    return render_template('timetable_batch.html',batches=batches)

@app.route('/even_batch_timetable/<int:sem>',methods=['GET'])
def even_batches_timetale(sem):
    batches = Batch.query.filter_by(odd_sem=sem).all()
    return render_template('timetable_batch.html',batches=batches)

@app.route('/classroom_timetable/<int:sem>',methods=['GET'])
def classroom_timetale(sem):
    classrooms = Classroom.query.all()
    return render_template('timetable_classroom.html',classrooms=classrooms,sem=sem)

@app.route('/lab_timetable/<int:sem>',methods=['GET'])
def lab_timetale(sem):
    labs = Lab.query.all()
    return render_template('timetable_lab.html',labs=labs,sem=sem)

@app.route('/get-timetable',methods=['GET'])
def get_timetable():
    return render_template('Odd_even_Timetable.html')

@app.route('/even_timetable/<int:sem>',methods=['GET','POST'])
def even_timetable(sem):
    if request.method == 'POST':
        batches=Batch.query.filter_by(odd_sem=sem).all()
        for batch in batches:
            db.session.query(Schedule).filter(
                Schedule.batch_id == batch.id,
                Schedule.semester == sem,
                Schedule.combined_course_id.is_(None)
            ).delete()
            db.session.commit()

            courses=Course.query.filter_by(batch_id=batch.id).all()
            for course in courses:
                add_this_in_schedule(course,sem)
    return render_template('even_timetable_home.html',sem=sem)

@app.route('/days/<int:sem>')
def show_days(sem):
    return render_template('days.html',sem=sem)

@app.route('/days-lab/<int:sem>')
def show_days_lab(sem):
    return render_template('days-lab.html',sem=sem)

@app.route('/day_lab/<int:day>/<int:sem>')
def day_pages_lab(day,sem):
    classrooms = Lab.query.all()
    schedules = Schedule.query.filter_by(day=day,semester=sem).all()

    occupied = {}
    for s in schedules:
        if s.lab_id is not None:
            if s.lab_id not in occupied:
                occupied[s.lab_id] = []
            occupied[s.lab_id].append(s.slot)
    slots = list(range(0, 10))
    slot_names = {
        0: "8:00 AM - 9:00 AM",
        1: "9:00 AM - 10:00 AM",
        2: "10:00 AM - 11:00 AM",
        3: "11:00 AM - 12:00 PM",
        4: "12:00 PM - 1:00 PM",
        5: "1:00 PM - 2:00 PM",
        6: "2:00 PM - 3:00 PM",
        7: "3:00 PM - 4:00 PM",
        8: "4:00 PM - 5:00 PM",
        9: "5:00 PM - 6:00 PM"
    }


    return render_template(
        'days_available_labs.html',
        day=day,
        classrooms=classrooms,
        occupied=occupied,
        slots=slots,
        slot_names=slot_names
    )

@app.route('/day_classroom/<int:day>/<int:sem>')
def day_pages(day,sem):
    classrooms = Classroom.query.all()
    schedules = Schedule.query.filter_by(day=day,semester=sem).all()

    occupied = {}
    for s in schedules:
        if s.classroom_id is not None:
            if s.classroom_id not in occupied:
                occupied[s.classroom_id] = []
            occupied[s.classroom_id].append(s.slot)
    slots = list(range(0, 10))
    slot_names = {
        0: "8:00 AM - 9:00 AM",
        1: "9:00 AM - 10:00 AM",
        2: "10:00 AM - 11:00 AM",
        3: "11:00 AM - 12:00 PM",
        4: "12:00 PM - 1:00 PM",
        5: "1:00 PM - 2:00 PM",
        6: "2:00 PM - 3:00 PM",
        7: "3:00 PM - 4:00 PM",
        8: "4:00 PM - 5:00 PM",
        9: "5:00 PM - 6:00 PM"
    }


    return render_template(
        'days_available_classrooms.html',
        day=day,
        classrooms=classrooms,
        occupied=occupied,
        slots=slots,
        slot_names=slot_names
    )

@app.route('/specific_batch_timetable/<int:id>', methods=['GET','POST'])
def specific_batch_timetable(id):
    if request.method == 'POST':
        print("HELLO")
        db.session.query(Schedule).filter(Schedule.batch_id == id).delete()
        courses=Course.query.filter_by(batch_id=id).all()
        batch=Batch.query.filter_by(id=id).first()
        for course in courses:
            add_this_in_schedule(course,batch.odd_sem)
    schedules = Schedule.query.filter_by(batch_id=id).all()
    timetable = [["-" for _ in range(10)] for _ in range(6)]

    print(len(schedules))
    for schedule in schedules:
        if schedule.elective_id>0:
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.filter_by(id=schedule.combined_course_id).first()
        if not course:
            continue
        professor = Professor.query.get(schedule.professor_id)
        classroom= Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        # Construct the timetable entry dynamically
        entry=""
        if schedule.tutorial and course:
            entry=f"{course.name} "
        else :
            if course:
                entry+=f"{course.name} "
            if professor:
                entry+=f"{professor.name}"
        if lab is not None:
            entry += f"(P) , {lab.name}"
        if classroom is not None:
            if schedule.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f"{classroom.name}"
        if schedule.divide_id>0:
            entry += f"A{schedule.divide_id}"
        entry += "\n"
        # Assign the constructed entry to the timetable
        timetable[schedule.day][schedule.slot] += entry
    
    for sch in schedules:
        if sch.elective_id==0:
            continue
        print(sch.id)
        course = Elective.query.get(sch.elective_course_id)
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry+=f"{sch.name} "
        else :
            entry += f"{sch.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"
        timetable[sch.day][sch.slot] += entry

    return render_template('show_timetable_batch.html', timetable=timetable,id=id)

@app.route('/specific_professor_timetable/<int:id>/<int:sem>', methods=['GET'])
def specific_professor_timetable(id,sem):
    schedules = Schedule.query.filter_by(professor_id=id,semester=sem).all()
    timetable = [["-" for _ in range(10)] for _ in range(6)]

    print(len(schedules))
    for schedule in schedules:
        if schedule.elective_id>0:
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.filter_by(id=schedule.combined_course_id)
        if not course:
            continue
        professor = Professor.query.get(schedule.professor_id)
        batch= Batch.query.get(schedule.batch_id)
        classroom= Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        # Construct the timetable entry dynamically
        entry = f""
        if course:
            entry+=f"{course.name} "
        if batch:
            entry+=f"{batch.name} "
        if schedule.divide_id>0:
            entry += f"A{schedule.divide_id}"
        entry += "\n"
        # Assign the constructed entry to the timetable
        timetable[schedule.day][schedule.slot] += entry
    
    for sch in schedules:
        if sch.elective_id==0:
            continue
        print(sch.id)
        course = Elective.query.get(sch.elective_course_id)
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry+=f"{sch.name} "
        else :
            entry += f"{sch.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"
        timetable[sch.day][sch.slot] += entry
    
    return render_template('show_timetable_professor.html', timetable=timetable,id=id)

@app.route('/specific_classroom_timetable/<int:id>/<int:sem>', methods=['GET'])
def specific_classroom_timetable(id,sem):
    flash("HELLO")
    schedules = Schedule.query.filter_by(classroom_id=id,semester=sem).all()
    timetable = [["-" for _ in range(10)] for _ in range(6)]

    print(len(schedules))
    for schedule in schedules:
        if schedule.elective_id>0:
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.filter_by(id=schedule.combined_course_id).first()
        if not course:
            continue
        professor = Professor.query.get(schedule.professor_id)
        batch= Batch.query.get(schedule.batch_id)
        classroom= Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        # Construct the timetable entry dynamically
        entry = f""
        if course:
            entry+=f"{course.name} "
        if batch:
            entry+=f"{batch.name} "
        if professor:
            entry+=f"{professor.name} "
        if schedule.divide_id>0:
            entry += f"A{schedule.divide_id}"
        entry += "\n"
        # Assign the constructed entry to the timetable
        timetable[schedule.day][schedule.slot] += entry
    
    batches=[]
    for sch in schedules:
        if sch.elective_id==0:
            continue
        if sch.classroom_id:
            if (sch.classroom_id,sch.day,sch.slot) in batches:
                continue
            else:
                batches.append((sch.classroom_id,sch.day,sch.slot))
        if sch.lab_id:
            if (sch.lab_id,sch.day,sch.slot) in batches:
                continue
            else:
                batches.append((sch.lab_id,sch.day,sch.slot))
        print(sch.id)
        course = Elective.query.get(sch.elective_course_id)
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry+=f"{sch.name} "
        else :
            entry += f"{sch.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"
        timetable[sch.day][sch.slot] += entry
    
    return render_template('show_timetable_classroom.html', timetable=timetable,id=id)

@app.route('/specific_lab_timetable/<int:id>/<int:sem>', methods=['GET'])
def specific_lab_timetable(id,sem):
    flash("HELLO")
    schedules = Schedule.query.filter_by(lab_id=id,semester=sem).all()
    timetable = [["-" for _ in range(10)] for _ in range(6)]

    print(len(schedules))
    for schedule in schedules:
        if schedule.elective_id>0:
            continue
        course = Course.query.get(schedule.course_id)
        if not course:
            course=CombinedCourse.query.filter_by(id=schedule.combined_course_id)
        if not course:
            continue
        professor = Professor.query.get(schedule.professor_id)
        batch= Batch.query.get(schedule.batch_id)
        classroom= Classroom.query.get(schedule.classroom_id)
        lab = Lab.query.get(schedule.lab_id)
        # Construct the timetable entry dynamically
        entry = f""
        if course:
            entry+=f"{course.name} "
        if batch:
            entry+=f"{batch.name} "
        if professor:
            entry+=f"{professor.name} "
        if schedule.divide_id>0:
            entry += f"A{schedule.divide_id}"
        entry += "\n"
        # Assign the constructed entry to the timetable
        timetable[schedule.day][schedule.slot] += entry
    
    batches=[]
    for sch in schedules:
        if sch.elective_id==0:
            continue
        if sch.classroom_id:
            if (sch.classroom_id,sch.day,sch.slot) in batches:
                continue
            else:
                batches.append((sch.classroom_id,sch.day,sch.slot))
        if sch.lab_id:
            if (sch.lab_id,sch.day,sch.slot) in batches:
                continue
            else:
                batches.append((sch.lab_id,sch.day,sch.slot))
        print(sch.id)
        course = Elective.query.get(sch.elective_course_id)
        prof = Professor.query.get(sch.professor_id)
        classroom = Classroom.query.get(sch.classroom_id)
        lab = Lab.query.get(sch.lab_id)
        entry=""
        if sch.tutorial:
            entry+=f"{sch.name} "
        else :
            entry += f"{sch.name} "
            if prof:
                entry+=f"({prof.name})"
        if classroom:
            if sch.tutorial:
                entry+= f" (T) "
            else:
                entry+= f" (L) "
            entry += f" [{classroom.name}]"
        if lab:
            entry += f"(P), [Lab: {lab.name}]"
        if sch.divide_id>0:
            entry += f"A{sch.divide_id}"
        entry += "\n"
        timetable[sch.day][sch.slot] += entry
    
    return render_template('show_timetable_lab.html', timetable=timetable,id=id)

if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(debug=True)